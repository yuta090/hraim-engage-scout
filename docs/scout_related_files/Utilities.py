import os,re
import pandas as pd
import openai
import requests
import tkinter as tk
import flet as ft
from tkinter import messagebox
import csv
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import platform
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

class OpenAIChatBot:
    def __init__(self, api_key):
        self.api_key = api_key
        openai.api_key = self.api_key

    def get_response(self, text, instruction):
        response = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": instruction},
                {"role": "user", "content": text}
            ]
        )
        return response.choices[0].message['content']
    
def check_engage_required_files():
    #  engageの更新に必要なファイルの存在チェック
    required_files = [
        "lib_data/station.csv",
        "lib_data/output.csv",
        "settings/engage_settings.xlsx"
    ]
    
    # 存在しないファイルのリスト
    missing_files = []
    
    # 各ファイルの存在を確認
    for file_path in required_files:
        if not os.path.exists(file_path):
            missing_files.append(file_path)
    
    # 存在しないファイルがある場合、エラーダイアログを表示
    if missing_files:
        error_message = "以下の必要なファイルが見つかりません:\n" + "\n".join(missing_files)
        root = tk.Tk()
        root.withdraw()  # メインウィンドウを非表示にする
        root.attributes('-topmost', True)  # ウィンドウを最前面に設定
        messagebox.showerror("ファイルエラー", error_message, parent=root)
        root.destroy()
        return False
    return True


def load_excel_to_df(filepath):
    """
    指定されたパスのExcelファイルを読み込み、データフレームとして返す。
    ファイルが存在しない場合はエラーを返す。

    Parameters:
    filepath (str): 読み込むExcelファイルのフルパス。

    Returns:
    pd.DataFrame: 読み込んだデータのデータフレーム。

    Raises:
    FileNotFoundError: ファイルが存在しない場合に発生。
    """
    if not os.path.exists(filepath):
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo("読み込みERROR", f"{filepath}がロードできませんでした。処理を終了します", parent=root)
        root.destroy()
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        #messagebox.showinfo("システム終了", "指定されたファイルが見つかりません。システムを終了します。", parent=root)
        root.destroy()
        #raise SystemExit("指定されたファイルが見つかりません。システムを終了します。")
        return False
    return pd.read_excel(filepath)

def find_column_index_by_name(sheet, column_name):
    for row in sheet.iter_rows(min_row=1, max_row=1):  # ヘッダー行だけを走査
        for cell in row:
            if cell.value == column_name:
                return cell.column  # 列番号を返す
            

def get_row_index_for_salary_type(salary_type):
    """指定された給与タイプに対応する行インデックスを返す。"""
    salary_type_to_row_index = {
        "年俸": 0,
        "月給": 2,
        "日給": 3,
        "時給": 4
    }
    return salary_type_to_row_index.get(salary_type)

# XPathとアクションを取得するためのヘルパー関数
def get_xpath_and_action(df, column_name, value, app):
    try:
        # 指定された値がDataFrameのカラムに存在するか確認します。
        if value in df.columns:
            # 存在する場合は、そのカラムの最初の行からXPathを取得します。
            xpath = df.at[0, value]
            # 同じく、二番目の行からアクションを取得します。
            action = df.at[1, value]
            # XPathまたはアクションがnullでないことを確認します。
            if pd.isnull(xpath) or pd.isnull(action):
                # どちらかがnullの場合は、ログにメッセージを残し、Noneを返します。
                app.log_add(f"カラム {value} のXPathまたはアクション値が無効です。")
                return None, None
            # XPathとアクションの値が有効な場合は、それらを返します。
            return xpath, action
        #カラム名が見つからない場合、engage-uploadの現在のカラムの値を元にカラム名を検索
        elif app.column_name in df.columns:
            xpath_value = df.at[0, app.column_name]
            action_value = df.at[1, app.column_name]
            if pd.isnull(xpath_value) or pd.isnull(action_value):
                app.log_add(f"カラム {value} は df の中に存在しません。")
                return None, None                
            return xpath_value, action_value
        else:
                return None, None
            
    except KeyError as e:
        # KeyErrorが発生した場合（指定されたキーがDataFrameに存在しない場合）、
        # ログにエラーメッセージを残し、Noneを返します。
        app.log_add(f"KeyError: {e} - カラム {column_name} は df の中に存在しません。")
        return None, None
    except IndexError as e:
        # IndexErrorが発生した場合（指定されたインデックスが範囲外の場合）、
        # ログにエラーメッセージを残し、Noneを返します。
        app.log_add(f"IndexError: {e} - 指定されたインデックスが範囲外です。")
        return None, None

# 与えられた行の特定のカラムに有効な値が存在するかどうかをチェックする関数
def has_valid_values(row, *columns):
    return any(row[col] for col in columns)


class AddressParser:
    def __init__(self, address):
        self.address = address

    def get_prefecture(self):
        """
        都道府県を取得するメソッド。
        """
        pattern = r'^(東京都|北海道|(?:京都|大阪)府|.{2,3}県)'
        match = re.match(pattern, self.address)
        return match.group(1) if match else None

    def get_city_ward_town_village(self):
        """
        市区町村を取得するメソッド。
        """
        pattern = r'^(東京都|北海道|(?:京都|大阪)府|.{2,3}県)(.*?[市区町村])'
        match = re.match(pattern, self.address)
        return match.group(2) if match else None

    def get_county(self):
        """
        郡を取得するメソッド。
        """
        pattern = r'.*?[市区町村](.*?[郡])'
        match = re.search(pattern, self.address)
        return match.group(1) if match else None

    def get_town_village(self):
        """
        町村を取得するメソッド。
        """
        pattern = r'.*?[市区町村](.*?[町村])'
        match = re.search(pattern, self.address)
        return match.group(1) if match else None

    def get_chome(self):
        """
        丁目を取得するメソッド。
        """
        pattern = r'.*?[町村](.*?丁目)'
        match = re.search(pattern, self.address)
        return match.group(1) if match else None

    def get_street_name(self):
        """
        地名を取得するメソッド。
        """
        pattern = r'.*?[町村].*?丁目(.*?)(\d+[-番丁目号]+.*)?$'
        match = re.search(pattern, self.address)
        return match.group(1).strip() if match and match.group(1) else None

    def get_street_number(self):
        """
        番地を取得するメソッド。
        """
        pattern = r'(?:.*?[町村].*?丁目)?(\d{1,3}(?:-\d{1,3}){0,2})'
        match = re.search(pattern, self.address)
        return match.group(1).strip() if match and match.group(1) else None


def get_zipcode_from_address(address):
    """住所から郵便番号を返す。2つのサイトから取得を試みる。"""
    try:
        # Web APIのURL
        api_url = "http://api.excelapi.org/post/zipcode"
        # リクエストを送信
        response = requests.get(api_url, params={'address': address})
        response.raise_for_status()  # エラーレスポンスをチェック
        # レスポンスをJSONとして読み込む
        data = response.text

        # 郵便番号を取得して返す
        if data:
            return data
        else:
            print("APIからのレスポンスに郵便番号が含まれていません。")
            raise Exception("APIからのレスポンスに郵便番号が含まれていません。")
    except Exception as e:
        print("APIからのレスポンスに郵便番号が含まれていません。")
        print("別のAPIを試します。")

        # 別のAPIのURL
        api_url = "https://zipcoda.net/api"
        # リクエストを送信
        response = requests.get(api_url, params={'address': address})
        response.raise_for_status()  # エラーレスポンスをチェック
        # レスポンスをJSONとして読み込む
        data = response.json()

        # 郵便番号を取得して返す
        if data['status'] == 200:
            return data['items'][0]['zipcode']
        else:
            print("別のAPIからも郵便番号が取得できませんでした。")
            return "郵便番号を取得できませんでした"

def are_all_true(row, *columns):
    """
    指定された行のカラムが全てTrueかどうかを確認する関数。

    Args:
    row (pd.Series): データフレームの行。
    *columns (str): 確認するカラム名の可変長引数。

    Returns:
    bool: 全てのカラムの値がTrueであればTrue、そうでなければFalse。
    """
    return all(row.get(column, False) for column in columns)

class LoggerApp:
    def __init__(self, root):
        self.root = root
        self.setup_gui()

    def setup_gui(self):
        """GUIのセットアップを行う"""
        self.root.title("ログ")
        self.root.attributes('-topmost', True)
        self.root.update()

        # ログを表示するテキストボックスの作成
        self.log_text = tk.Text(self.root, height=10, state=tk.DISABLED)
        self.log_text.pack()

        # log_message関数を初回呼び出し
        self.log_message("プログラムが開始しました。")

    def log_message(self, message):
        """ログメッセージをテキストボックスに追加するメソッド"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.config(state=tk.DISABLED)
        # テキストボックスを最新のログにスクロール
        self.log_text.see(tk.END)
        # GUIを更新するためのコードはコメントアウトされています
        self.root.update()
        # ここでは再帰的な呼び出しを行わない


class TextProcessor:
    @staticmethod
    def format_time(time_str):
        """時間をHH:MM形式に整形し、時間と分に分割する。"""
        # 時間が一桁の場合、先頭に0を追加
        formatted_time = time_str.zfill(5)
        # 時間と分に分割
        hours, minutes = formatted_time.split(':')
        return hours, minutes

    @staticmethod
    def remove_station_name(text):
        """
        入力されたテキストから「XX駅」とその前のスペースを含めて削除する関数。

        Args:
        text (str): 処理するテキスト

        Returns:
        str: 処理後のテキスト
        """
        return text.replace('駅', '')

    # station.csvから該当する駅名の住所を検索する関数
    @staticmethod
    def search_address_in_csv(station_name, csv_file_path):
        with open(csv_file_path, mode='r', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row in csv_reader:
                if station_name in row:
                    # 7番目の列は郵便番号、8番目の列は住所）
                    return row[8], row[7]
        return None

    @staticmethod
    def remove_symbols_and_spaces(text):
        """
        テキストから半角全角の記号、半角全角のスペースを除去する関数。

        Args:
        text (str): 処理するテキスト

        Returns:
        str: 処理後のテキスト
        """
        cleaned_text = re.sub(r'[!-~]|[︰-＠]|[、-〜]|[　\s]|・', '', text)
        return cleaned_text
    
    def convert_columns_to_int64(self,df, columns_list):
        """
        指定されたカラムのデータ型をInt64に変換する関数。
        
        :param df: データ型を変換するDataFrame
        :param columns_list: データ型を変換するカラム名のリスト
        :return: データ型が変換されたDataFrame
        """
        for column in columns_list:
            if column in df.columns:
                df[column] = df[column].astype(pd.Int64Dtype(), errors='ignore')
        return df

def show_debug_message(index, column_name):
    """
    デバッグモードのメッセージボックスを最前面に表示する関数。
    """
    # Tkウィンドウを作成
    root = tk.Tk()
    root.withdraw()  # メインウィンドウを非表示にする
    root.attributes('-topmost', True)  # ウィンドウを最前面に設定

    # メッセージボックスを表示
    messagebox.showinfo("デバッグモードON", f"{index + 1}件目\n {column_name}", parent=root)

    # Tkウィンドウを破棄
    root.destroy()
    
print("")# 辞書クラス作成
class Lookup:
    def __init__(self, list1, list2):
        # 辞書をクラスの属性として初期化
        self.lookup_dict = dict(zip(list1, list2))

    def get_value(self, key):
        # 与えられたキーに対応する値を返す
        return self.lookup_dict.get(key, None)


class ImageChecker:
    def __init__(self, company_dir,app, driver=None):
        """
        ImageChecker クラスのコンストラクタ。
        
        :param company_dir: 画像ファイルが格納されているディレクトリのパス。
        :param driver: (オプション) WebDriverのインスタンス。デフォルトはNone。
        """
        self.company_dir = company_dir
        self.app = app  # appをインスタンス変数として保存
        self.driver = driver

    def check_images_existence(self):
        self.app.log_add("images.xlsxの存在チェック…")
        """会社名/images.xlsxを読み込み、画像の存在チェックを行う"""
        xlsx_path = os.path.join(self.company_dir, 'images.xlsx')
        if not os.path.exists(xlsx_path):
            self.app.log_add("images.xlsxが見つかりません。")
            return None

        df = pd.read_excel(xlsx_path)
        if '画像名' not in df.columns or 'アンカーテキスト' not in df.columns:
            self.app.log_add("必要な列が存在しません。")
            return None

        self.app.log_add("images.xlsxの存在チェック…OK")
        self.app.log_add("images.xlsxの画像データを確認中…")
        missing_images = []
        for image_name in df['画像名']:
            image_path = os.path.join(self.company_dir, 'images', image_name)
            if not os.path.exists(image_path):
                missing_images.append(image_name)

        if missing_images:
            root = tk.Tk()
            root.withdraw()  # Tkのルートウィンドウを表示しない
            messagebox.showerror("エラー", "以下の画像がimagesフォルダに見つかりません:\n" + "\n".join(missing_images))
            root.destroy()
            return None

        self.app.log_add("画像チェック…OK")
        return df[['画像名', 'アンカーテキスト']]

    def upload_file(self, file_path, element_type, element_value):
        """
        ファイルをアップロードする。このメソッドを使用するには、コンストラクタで
        WebDriverのインスタンスを渡す必要がある。
        """
        if not self.driver:
            raise ValueError("WebDriverのインスタンスが設定されていません。")

        # 相対パスを絶対パスに変換（ファイルパスが相対パスの場合）
        absolute_file_path = os.path.abspath(file_path)
        # 指定されたエレメントを見つけてファイルパスを送信
        self.driver.find_element(element_type, element_value).send_keys(absolute_file_path)
        
def is_file_open(file_path):
    if platform.system() == 'Darwin':  # macOS の場合
        return is_file_open_mac(file_path)
    else:  # Windows または他のプラットフォームの場合
        return is_file_open_default(file_path)

def is_file_open_default(file_path):
    try:
        # 読み取り専用モードでファイルを開く
        with open(file_path, 'r+'):
            pass
    except IOError:
        return True
    return False

def is_file_open_mac(file_path):
    if not os.path.isfile(file_path):
        return False

    try:
        with open(file_path, 'r+'):
            pass
    except IOError:
        return True

    try:
        with open(file_path, 'r+') as file:
            try:
                import fcntl
                fcntl.flock(file, fcntl.LOCK_EX | fcntl.LOCK_NB)
                fcntl.flock(file, fcntl.LOCK_UN)
            except IOError:
                return True
    except IOError:
        return True

    return False



def write_df_to_excel_with_validation(file_path, df_download_data, validation_column="更新フラグ"):
    """
    指定されたExcelファイルにDataFrameの内容を書き込み、特定の列にデータバリデーションを追加する関数。

    Parameters:
    file_path (str): Excelファイルのパス。
    df_download_data (pd.DataFrame): 書き込むデータフレーム。
    validation_column (str): データバリデーションを追加する列名。デフォルトは"更新フラグ"。
    """
    # EXCELファイルの読み込み
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # DataFrameをEXCELに書き込む
    for r_idx, row in enumerate(dataframe_to_rows(df_download_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 更新フラグカラムのインデックスを取得
    update_flag_col_idx = df_download_data.columns.get_loc(validation_column) + 1

    # チェックボックスのデータバリデーションを作成
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=True)
    ws.add_data_validation(dv)

    # チェックボックスを更新フラグカラムに適用
    for row in range(2, ws.max_row + 1):  # 1行目はヘッダーなので2行目から開始
        cell = ws.cell(row=row, column=update_flag_col_idx)
        dv.add(cell)

    # EXCELファイルを保存
    wb.save(file_path)

# 使用例
# write_df_to_excel_with_validation("path/to/excel/file.xlsx", df_download_data)

def show_error_message(page, title, message):
    dialog = ft.AlertDialog(
        title=ft.Text(title),
        content=ft.Text(message),
        actions=[
            ft.TextButton("OK", on_click=lambda e: page.dialog.close())
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    page.dialog = dialog
    dialog.open = True
    page.update()

def show_error_message(page, title, message):
    dialog = ft.AlertDialog(
        title=ft.Text(title),
        content=ft.Text(message),
        actions=[
            ft.TextButton("OK", on_click=lambda e: page.dialog.close())
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    page.dialog = dialog
    dialog.open = True
    page.update_async()

def show_message_box(page, title, message, box_type="error"):
    result = None

    def on_ok(e):
        nonlocal result
        result = True
        page.dialog.open = False
        page.update()
    
    def on_yes(e):
        nonlocal result
        result = True
        page.dialog.open = False
        page.update()

    def on_no(e):
        nonlocal result
        result = False
        page.dialog.open = False
        page.update()

    if box_type == "error":
        dialog = ft.AlertDialog(
            title=ft.Text(title),
            content=ft.Text(message),
            actions=[
                ft.TextButton("OK", on_click=on_ok)
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
    elif box_type == "askyesno":
        dialog = ft.AlertDialog(
            title=ft.Text(title),
            content=ft.Text(message),
            actions=[
                ft.TextButton("Yes", on_click=on_yes),
                ft.TextButton("No", on_click=on_no)
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
    page.dialog = dialog
    dialog.open = True
    page.update()

    # Wait until the dialog is closed
    while dialog.open:
        time.sleep(0.1)
   
    return result

class ScoutConfig:
    def __init__(self):
        # デフォルト設定
        self.min_age = 21
        self.max_age = 60
        self.target_prefecture = ""
        self.keywords = ["プロフィール"]
        
        # 設定ファイルからの読み込み
        self.load_settings()
    
    def load_settings(self):
        """設定ファイルから設定を読み込む"""
        try:
            settings_df = load_excel_to_df('settings/scout_settings.xlsx')
            if settings_df is not None:
                if '最小年齢' in settings_df:
                    self.min_age = int(settings_df['最小年齢'].iloc[0])
                if '対象都道府県' in settings_df:
                    self.target_prefecture = settings_df['対象都道府県'].iloc[0]
                if 'キーワード' in settings_df:
                    keywords = settings_df['キーワード'].iloc[0]
                    if isinstance(keywords, str):
                        self.keywords = [kw.strip() for kw in keywords.split(',')]
        except Exception as e:
            print(f"設定ファイルの読み込みに失敗しました: {e}")

    def update_max_age(self, age):
        """タブからの入力値で最大年齢を更新"""
        try:
            self.max_age = int(age)
            return True
        except ValueError:
            return False

    def update_age_range(self, min_age, max_age):
        """年齢範囲を更新"""
        try:
            min_val = int(min_age) if min_age is not None else self.min_age
            max_val = int(max_age) if max_age is not None else self.max_age
            
            if min_val > max_val:
                # 最小値が最大値より大きい場合は入れ替え
                min_val, max_val = max_val, min_val
                
            self.min_age = min_val
            self.max_age = max_val
            return True
        except (ValueError, TypeError):
            return False

    def validate_age(self, age):
        """年齢が有効範囲内かチェック。値に単位の文字列があれば削除"""
        try:
            # 年齢から単位の文字列を削除
            if isinstance(age, str):
                age = age.replace("歳", "").replace("才", "")
            age_val = int(age)
            return self.min_age <= age_val <= self.max_age
        except (ValueError, TypeError):
            return False

    def check_keywords(self, text):
        """キーワードの存在チェック"""
        return any(keyword in text for keyword in self.keywords)
    
