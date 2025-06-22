# import atexit
import os
import pandas as pd
import threading

import time
import random

from openpyxl import load_workbook

import tkinter as tk
from tkinter import messagebox

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC

from lib.Utilities import OpenAIChatBot,LoggerApp,ImageChecker,TextProcessor,are_all_true
from lib.Utilities import get_zipcode_from_address,load_excel_to_df,get_row_index_for_salary_type,find_column_index_by_name
from lib.Utilities import get_xpath_and_action,AddressParser,check_engage_required_files,is_file_open,show_error_message,show_message_box,ScoutConfig
from lib.browser import BrowserManager
from lib.browser_engage import EngageSeleniumAutomation,engage_auth_check
from lib.mail import *  #メール関係の処理実装
from lib.browser_airwork import *
from urllib.parse import urlparse, parse_qs

import flet as ft
from flet import Page, Column, Row, Text, ElevatedButton, NavigationRail, NavigationRailDestination, TextField, icons, colors, alignment
from collections import deque

import socket,logging,time
from functools import reduce
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

#アプリのバージョン指定
global PROGRAM_VER
PROGRAM_VER = "0.194"

def main(page: ft.Page):
    #logフォルダがカレントディレクトリになければ作成する。
    if not os.path.exists('log'):
        os.makedirs('log')

    # ロガーの設定
    logging.basicConfig(
        filename=f'log/{datetime.today().strftime("%Y-%m-%d")}_app.log',
        filemode='a',
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        level=logging.ERROR
    )

    # コンソールにもログを出力するためのハンドラを追加
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

    # ルートロガーにハンドラを追加
    logging.getLogger().addHandler(console_handler)

    # メインファイルのロガーを作成
    logger = logging.getLogger(__name__)

    #初期設定各種
    app = MyApp(page)
    
    # 新しい方法でウィンドウを前面に表示
    page.window.to_front()
    
    # engage認証情報の取得と初期化
    engage_id, engage_pass, company_name, engage_id_list, engage_pass_list, company_name_list = engage_auth_check(app, PROGRAM_VER)
    
     # 認証データの初期化
    app.initialize_auth_data(
        engage_id,
        engage_pass,
        company_name,
        engage_id_list,
        engage_pass_list,
        company_name_list,

    )


    #engageの設定ファイルを読み込み
    df_setthing = None
    engage_settings = "settings/engage_settings.xlsx"
    df_setthing = load_excel_to_df(engage_settings)

    # 必須ファイルチェックを実行
    if not check_engage_required_files():
        raise SystemExit("必要なファイルが足りないため、プログラムを終了します。")

    if '文字数オーバー時' in df_setthing:
        app.setting_text_over = df_setthing['文字数オーバー時'].iloc[0]
    if '住所自動入力ボタン' in df_setthing:
        app.setting_addr_auto_insert_button = df_setthing['住所自動入力ボタン'].iloc[0]
    if '職種カテゴリが見つからない場合' in df_setthing:
        app.setting_job_category_not_found = df_setthing['職種カテゴリが見つからない場合'].iloc[0]

    if 'デバッグモード' in df_setthing:
        app.setting_debug_mode = df_setthing['デバッグモード'].iloc[0]
    app.log_add("設定ファイル読み込み…OK")

    # 必要な設定項目のリスト
    required_settings = [
        'setting_text_over', 
        'setting_addr_auto_insert_button', 
        'setting_job_category_not_found', 
        'setting_debug_mode'
    ]

    # 空の設定項目を格納するリスト
    empty_settings = []

    # 各設定項目をチェックし、空の場合はリストに追加
    for setting in required_settings:
        if not getattr(app, setting, None):
            empty_settings.append(setting)

    # 空の設定項目がある場合、メッセージボックスで表示し、処理を中断
    if empty_settings:
        root = tk.Tk()
        root.withdraw()  # メインウィンドウを表示しない
        messagebox.showerror("設定エラー", f"以下の設定項目が空です: {', '.join(empty_settings)}")
        raise SystemExit("設定項目が空のため、プログラムを終了します。")

def engage_upload(app,e=None,login=None,up_filepath=None):
    tp = TextProcessor()

    #外部からの呼び出しの際に、別のアップロード用ファイルを指定できるように修正2024.6.20
    if up_filepath is None:
        # アップロードデフォルトのファイルパスを指定
        file_path = os.path.join(app.company_dir, 'engage-upload.xlsx')
    else:
        #デフォルトではないファイルパス
        file_path = up_filepath
        
    if os.path.exists(file_path):
        # ファイルが存在する場合の処理をここに記述
        app.log_add(f"{file_path}…ロードOK")
        # ここに具体的な処理を実装する
        try:
            # ファイルを追記モードで開く試み
            with open(file_path, 'a', encoding='utf-8') as file:
                # ファイルが問題なく開けるようであればファイルをDFと、openpyxlに読み込み
                df = load_excel_to_df(file_path)
                #openpyxlライブラリを使用してファイルを開く。保存時の各種フォーマット解除対策。
                book = load_workbook(file_path)

                # ここでデータフレーム `df` に対する処理を続ける
                pass
        except IOError as e:
            # ファイルが開けなかった場合、エラーをログに記録し、ユーザーに通知する
            app.show_message(e,"error", f"ファイル {file_path} が既に開かれています。閉じてから再度実行してください")
            return "プログラムを終了します。"

    else:
        # ファイルが存在しない場合の処理をここに記述
        app.show_message(e,"error", f"{app.company_dir}engage-upload.xlsxファイルが存在しません。")
        return "プログラムを終了します。"

    #openpyxlで操作するシートを指定
    sheet_name = '入力リスト'
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet = book[sheet_name]
    # 列名から列番号を取得する関数を使用
    engage_column_index = find_column_index_by_name(sheet, 'engage')
    engage_update_column_index = find_column_index_by_name(sheet, 'engage_update')


    # Engage自動入力処理開始
    # Engageログイン処理
    bm = BrowserManager()
    selenium_en = EngageSeleniumAutomation(bm.driver, app)

    #ログイン後最初に開くページのURL
    default_url = "https://en-gage.net/company/job/regist/form/?via=job"
    
    if login == "update":
        default_url = "https://en-gage.net/company/job"
    login_success = selenium_en.login(
        url=default_url,
        login_id=app.engage_id,
        password=app.engage_pass
    )
    if not login_success:
        return

    #各カラムの更新に使用するXpath一覧をロード
    df_EngageXpath = load_excel_to_df('settings/engage-xpath.xlsx')
    #print(df_EngageXpath.columns)

    # 0から始まる数字を数値型から文字列型として読み込むよう指定。
    df_engage_upload = pd.read_excel(file_path,dtype={"想定勤務：開始時間": str, "想定勤務：開始時間": str, "想定勤務：開始分": str, "想定勤務：終了時間": str, "想定勤務：終了分": str})

    #ダウンロードデータを更新する際は、■更新フラグが立っているレコードを抽出する。
    if up_filepath is not None:
        if '■更新フラグ' in df_engage_upload.columns:
            app.log_add("■更新フラグ カラムが存在しません。処理を中断します。")
            df_engage_upload = df_engage_upload[df_engage_upload['■更新フラグ'] == 1]        

    #浮動小数点と誤認されやすいカラムリスト一覧
    columns_to_convert = ['仕事No.','郵便番号','以降の住所','給与（最低額）','給与（最高額）','想定年収（最低額）','想定年収（最高額）','年収例_1','年収例_2','年収例_3',]

    # 浮動小数点と誤認されるカラムの変換処理
    df_engage_upload = tp.convert_columns_to_int64(df_engage_upload, columns_to_convert)
    column_name = ""


    #住所の自動入力ボタンのXPATH
    xpath_auto_addr_input_button = '//*[@id="jobMakeForm"]/div[1]/div[2]/dl/dd[4]/div/div/dl/dd[3]/div/a' 

    #原稿更新が成功件数カウント用
    update_colum_count = 0
    for index, row in df_engage_upload.iterrows():
        app.log_add(str(index + 1) + "件目処理中…")
        if up_filepath is not None:
            #原稿のコピー処理
            # 求人IDカラムの内容を取得
            work_id = row['work_id']
            if isinstance(work_id, float):
                work_id = str(int(work_id))
            else:
                work_id = str(work_id)

            # 原稿をコピーするURLを開く
            url_template = "https://en-gage.net/company/job/regist/form?engagejobList={work_id}&via_joblist&PK={app.pk_value}"
            generated_url = url_template.format(work_id=work_id, app=app)            

            bm.driver.get(generated_url)
            # 最大10秒間待機し、指定されたXPathの要素に特定のテキストが表示されるのを待つ
            WebDriverWait(bm.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
            )
            
        excel_row_index = index + 2  # Excelの行は1から始まるため、適宜調整（ヘッダー行がある場合は+2）
        if pd.notnull(row["engage"]) and row["engage"]:
            app.log_add(f"更新済みまたはデータなし_{row['表示用職種名']}： {row['市区町村']}")
            continue
        # 列名のリストを取得k
        column_names = list(row.index)
        # 例外処理カラムリスト
        list_form = [
            '雇用形態','派遣:選択', '職種カテゴリー', '都道府県','勤務先区分','求人区分', '給与タイプ', '支払方法', 
            'みなし残業代', '勤務時間', '想定勤務：開始時間', '想定勤務：開始分', '想定勤務：終了時間', 
            '想定勤務：終了分', '最終学歴', '募集職種の経験有無', '休みの取り方'
        ]

        # Seriesオブジェクトの列名を全て取得し、処理を行う
        for column_name in column_names:
            #以降の住所が空だった場合の例外処理
            #最寄り駅の値が空でなければ郵便版を宇を取得し、住所の自動入力をクリックし、番地情報があれば削除
            if column_name == "以降の住所" and pd.isnull(row[column_name]):
                app.log_add("以降の住所が空白のため駅の住所を挿入")
                #郵便番号と以降の住所のXpath。
                xpath_zip = '//*[@id="work_office_zip_code[0]"]'
                xpath_street = '//*[@id="other_address[0]"]'
                selenium_en.fill_address_from_station(row, tp, xpath_zip, xpath_auto_addr_input_button, xpath_street)

            #マップ：地図のクリック
            if column_name == "最寄り駅":
                selenium_en.click_map(column_name)

            #現在のカラム名をクラス変数に格納
            app.column_name = column_name

            # 変数初期化：xpath保存用変数、Seleniumのアクション保存用変数。
            xpath_value = None
            action_value = None

            #現在のカラムに値が無ければ次の行へ
            if pd.isnull(row[column_name]) or row[column_name] in [None, '', False]:
                continue

            #例外処理リストの判定。TRUEの場合
            if column_name in list_form:
                # データフレームにセレクトボックスの列が存在すれば、xpathと実行するactionを取得。
                # カラムの値がnullまたは無効な値でないことを確認
                if pd.isnull(row[column_name]) or row[column_name] in [None, '', False]:
                    continue

                # XPathとアクションを取得
                xpath_value, action_value = get_xpath_and_action(df_EngageXpath, column_name, row[column_name],app)
                if xpath_value is None or action_value is None:
                    app.log_add(f"カラム {column_name} のXPathまたはアクション値が無効です。")
                    continue

                # 職種カテゴリーの取得とクリック処理
                if column_name == "職種カテゴリー":
                    if not selenium_en.search_and_click_elements_job_category(bm, row[column_name]):
                        app.log_add(f"職種カテゴリー_NG")
                        continue
                    else:
                        continue

                #給与入力の処理をする際、年俸や月給によって入力するXPATHが違うので、動的に取得するよう変更
                if column_name == "給与":
                    continue
            else:
                # 現在のカラム名がengage-xpath.xlsxに同名のカラムが存在するか確認
                if column_name not in df_EngageXpath.columns: 
                    logging.info(f"[{column_name}] が engage-xpath.xlsx に存在しません。")
                    continue

                action_value = df_EngageXpath.at[1, column_name]

                #給与入力の処理をする際、年俸や月給によって入力するXPATHが違うので、動的に取得するよう変更
                if column_name in ["給与（最低額）", "給与（最高額）"]:
                    salary_no = get_row_index_for_salary_type(row["給与タイプ"])
                    xpath_value = df_EngageXpath.at[salary_no, column_name]
                else:
                    # 最初と二番目の行の値が空でないか、nan、None、FALSEでないか確認
                    xpath_value = df_EngageXpath.at[0, column_name]

                #画像のアップロード処理
                if "画像ファイル" in column_name:
                    img_last_character = ""
                    img_last_character = column_name[-1]

                    #原稿アップロード時に画像ファイルがすでにある場合に、1枚目の画像ファイルだけXPATHを変更
                    # cssセレクタを使用して要素を取得
                    element = bm.driver.find_element(By.CSS_SELECTOR, 'div.fileInput.photo.js_editImage > label')
                    # 要素のinnerHTMLを取得
                    inner_html = element.get_attribute('innerHTML')
                    # innerHTMLに'svg'が含まれているか判定
                    if column_name == "画像ファイル_1" :
                        if 'svg' in inner_html:
                            print("1枚目は空です")
                        else:
                            xpath_value = '//*[@id="pictureArea"]/a[1]/div/div[3]'
                            print("1枚目に画像が含まれています")

                    if selenium_en.hover_over_element(xpath_value, img_last_character, row[column_name]):
                        continue

                #画像コメント_1が、更新時にはXpathが変わるので例外対応
                if column_name == "画像コメント_1" :
                    if up_filepath:
                        xpath_value = '//*[@id="pictureArea"]/a[1]/div/div[4]/textarea[2]'

                if "休日休暇" in column_name:
                    print("休日休暇")

                # 特定のチェックボックスがONの場合は各各種の備考の入力をスキップする。
                # 処理をスキップする条件
                skip_conditions = {
                    "勤務地：備考": ('駅から徒歩5分以内', '転勤なし', 'テレワーク・在宅OK'),
                    "休日休暇": ('夏季休暇', '年間休日120日以上', '年末年始休暇'),
                    "待遇・福利厚生": ('雇用保険', '労災保険', '厚生年金', '健康保険', '交通費支給あり', '資格取得支援・手当あり', '寮・社宅・住宅手当あり', '育児支援・託児所あり', 'U・Iターン支援あり', '時短勤務制度あり', '日払い・週払い・即日払いOK', '服装自由')
                }
                # 条件に一致する場合は次のカラムへスキップ
                #if column_name in skip_conditions and has_valid_values(row, *skip_conditions[column_name]):
                #    continue            # Seleniumを使用してアクションを実行

            #データ挿入処理
            try:
                if action_value == 'click':
                    selenium_en.fill_web_form('XPATH', xpath_value, action_value)
                elif action_value == 'send_keys':
                    #文字数の入力制限対策処理。engage-xpath.xlsxから各カラムの最大入力値を取得しfill_web_form関数へ渡す。
                    max_input_length = None
                    try:
                        # DataFrameから特定の行と列の値を取得します。
                        # ここでは8行目とcolumn_name列の値を取得しています。
                        # pd.isnull() 関数を使用して値がnullでないことを確認し、nullでなければその値を max_input_length に設定します。
                        max_input_length = df_EngageXpath.at[8, column_name] if not pd.isnull(df_EngageXpath.at[8, column_name]) else None
                    except KeyError:
                        # KeyErrorが発生した場合、max_input_length を None に設定します。
                        # これは、指定された行または列がDataFrameに存在しない場合に発生します。
                        max_input_length = None
                    # Seleniumを使用してWebフォームを埋める関数を呼び出します。
                    # ここではXPath、アクションタイプ、入力値、最大入力長を引数として渡しています。
                    selenium_en.fill_web_form('XPATH', xpath_value, action_value, row[column_name], max_input_length)
                elif action_value == 'select_by_visible_text':
                    selenium_en.fill_web_form('XPATH', xpath_value, action_value, row[column_name])
                elif action_value == 'select_by_value':
                    selenium_en.fill_web_form('XPATH', xpath_value, action_value, row[column_name])
                # 必要に応じてさらにアクションを追加
            except Exception as e:
                app.log_add(f"{column_name} エラーが発生しました: {e}")
                            #都道府県が未入力だった場合
            #  郵便番号入力完了後に、自動入力ボタンをクリックするかの例外処理
            if column_name  == "郵便番号" :
                if app.setting_addr_auto_insert_button == "強制クリック":
                    selenium_en.fill_web_form('XPATH', xpath_auto_addr_input_button, 'click')
                elif app.setting_addr_auto_insert_button == "何もしない":
                    #この処理は何もせずに次の列へ
                    continue
                elif app.setting_addr_auto_insert_button == "都道府県未入力時クリック":
                    #この値の場合は、都道府県の値をチェックし、NULLならクリック処理を行う。
                    if pd.isnull(row['都道府県']):
                        selenium_en.fill_web_form('XPATH', xpath_auto_addr_input_button, 'click')
                else:
                    #選択し以外の場合は何もせず次の列へ
                    continue

        #求人掲載にあたっての確認事項のチェック処理
        checkboxes_xpath = "/html/body/div[1]/form/div[2]/div/div[2]/div[2]/dl//input[@type='checkbox']"
        try:
            # チェックボックス要素を全て取得
            checkboxes = WebDriverWait(bm.driver, 3).until(
                EC.presence_of_all_elements_located((By.XPATH, checkboxes_xpath))
            )
            # 各チェックボックスにチェックを入れる
            for checkbox in checkboxes:
                try:
                    #チェックボックス最後の値が隠し要素のため、そのタイミングで処理を抜ける
                    if checkbox == checkboxes:
                        break
                    # 強制的にクリックを試みる
                    bm.driver.execute_script("arguments[0].click();", checkbox)
                except Exception as e:
                    # エラーメッセージを出力するが、ループは続行する
                    print(f"チェックボックスをクリック中にエラーが発生しました: {e}")

        except TimeoutException:
            print('チェックボックスが見つかりませんでした。')
        except Exception as e:
            app.log_add("チェックボックスの取得中にエラーが発生しました")
            print(f"チェックボックスの取得中にエラーが発生しました: {e}")

        #終了処理。ステータスが入力内容を保存なら、原稿を保存して次の原稿作成へ。
        #新規追加なら原稿原稿の追加処理へ。
        if row['ステータス'] == '入力内容を保存':
            # Aの処理
            xpath_value = '//*[@id="jobMakeFunction"]/div/div[2]/div/a[1]'
            action_value = 'click'
            selenium_en.fill_web_form('XPATH', xpath_value, action_value)
            app.log_add("原稿を保存しました。")

        elif row['ステータス'] == '新規追加':
            # Bの処理
            app.log_add("原稿新規追加")

                        # スクリーンショットを取得し保存
            filename = os.path.join(os.path.dirname(os.path.abspath(__file__)),"screenshot_mobile.png")
            bm.driver.save_screenshot(filename)

            # 内容を確認するをクリック処理
            try:
                element = bm.driver.find_element(By.XPATH, "/html/body/div[1]/form/div[3]/a")
                element.click()
                print("要素をクリックしました.")

                # 次のページにある要素が表示されるのを待つ
                new_element = WebDriverWait(bm.driver, 4).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/header/div/div[3]/div/ul/li"))
                )
                # 次のページにある要素を操作する処理をここに記述する

                print("入力エラーがありました。処理を中断します")

                raise SystemExit

            except NoSuchElementException:
                print("指定されたXPathの要素が見つかりませんでした。処理を中断します。")
            except TimeoutException:
                print("次のページにある要素が見つかりませんでした、またはタイムアウトしました。処理を継続します。")
            except Exception as e:
                print(f"エラーが発生しました: {e}")

            try:
                # 要素をクリック
                element = bm.driver.find_element(By.XPATH, "/html/body/div[2]/div/a[2]")
                element.click()

                # 次のページが完全に表示されるのを待つ
                WebDriverWait(bm.driver, 10).until(
                    lambda driver: driver.execute_script('return document.readyState') == 'complete'
                )

            except NoSuchElementException:
                print("要素が見つかりませんでした。")
            except TimeoutException:
                print("次のページが完全に表示されませんでした。")
                print(f"エラーが発生しました: {e}")
        app.log_add(f"更新完了_{row['表示用職種名']}： {row['市区町村']}")
        engage_cell = sheet.cell(row=excel_row_index, column=engage_column_index)
        engage_update_cell = sheet.cell(row=excel_row_index, column=engage_update_column_index)

        engage_cell.value = 1
        engage_update_cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if app.setting_debug_mode is not None and app.setting_debug_mode == "ON":
            # デバッグモードがONの時にはメッセージボックスを表示して、更新前の状態を確認できるようにする。
            root = tk.Tk()
            root.attributes("-topmost", True)
            root.withdraw()
            messagebox.showinfo("デバッグモードON", f"{index +1 }件目\n {column_name}", parent=root)
            root.destroy()
        # ファイルを保存
        book.save(file_path)

        #原稿が更新された数をカウントアップ
        update_colum_count += 1
        # ログインページにアクセス
        bm.driver.get('https://en-gage.net/company/job/regist/form/?via=job')
        # 最大10秒間待機し、指定されたXPathの要素に特定のテキストが表示されるのを待つ　
        WebDriverWait(bm.driver, 10).until(
            EC.text_to_be_present_in_element(
                (By.XPATH, '/html/body/header/div/div[2]/div'), '求人の新規作成'
            )
        )
 
    # メールアドレスのリストを取得
    email_addresses = []
    try:
        # mail.csvファイルを読み込む
        df_emails = pd.read_csv('settings/mail.csv')
        # 'email'列からメールアドレスをリストに追加
        email_addresses = df_emails['email'].tolist()
    except Exception as e:
        print(f"メールアドレスの読み込み中にエラーが発生しました: {e}")
    # MailSender クラスのインスタンスを作成
    mailer = MailSender('sv13425.xserver.jp', 587, 'report@recruit-robo.info', '4eWEjri233oqwjDFMK')

    subject = "engage自動更新が完了しました"
    if login == "update":
        subject = "engage原稿コピーが完了しました"

    pc_name = socket.gethostname()
    print(f"PCの名前: {pc_name}")
    # メールの内容を設定
    mailer.set_message(
        from_addr='report@recruit-robo.info',
        to_addrs=email_addresses,
        subject=subject,
        body=f"{app.company_dir}：({update_colum_count}件)\n更新PC:{pc_name}",
        cc=[''],
        bcc=['']
    )

    # 添付ファイルを追加：ファイルが見つからなければ処理をスキップ
    #filename=f'log/{datetime.today().strftime("%Y-%m-%d")}_app.log',  # ログを保存する
    #if os.path.exists(filename):
    #    mailer.add_attachment(filename)


    # メールを送信
    mailer.send()

    app.log_add(f"{app.company_dir}の更新が完了：({update_colum_count}件)")

def engage_download(app, error_check=None ,e=None,Login=None):
    # リストを再初期化
    engage_id, engage_pass, company_name, engage_id_list, engage_pass_list, company_name_list = engage_auth_check(app, PROGRAM_VER, e)
    # 初期化したリストをアプリの属性に設定
    app.engage_id_list = engage_id_list
    app.engage_pass_list = engage_pass_list
    app.company_dir_list = company_name_list
    
    tp = TextProcessor()
    bm = None
    email_body = ""  # email_bodyを初期化
    
    #account.xlsxを基にループ。error_checkがNoneならループは1回のみで終了
    while app.engage_id_list:
        current_engage_id = app.engage_id_list.popleft()
        current_engage_pass = app.engage_pass_list.popleft()
        current_company_name = app.company_dir_list.popleft()
        app.engage_id = current_engage_id
        app.engage_pass = current_engage_pass
        app.company_dir = f"company_data/{current_company_name}/"

        # 既にBMでブラウザを開いている場合はブラウザを閉じる
        if hasattr(bm, 'driver') and bm.driver is not None:
            try:
                bm.driver.quit()
                app.log_add("既存のブラウザを閉じました。")
            except Exception as e:
                app.log_add(f"既存のブラウザを閉じる際にエラーが発生しました: {e}")
            finally:
                bm.driver = None
        bm = BrowserManager()
        selenium_en = EngageSeleniumAutomation(bm.driver, app)

        #エラーチェックmodeの時は、リスト取得を専用のEXCELファイルに出力する。
        if error_check is None:
            file_path ='engage-download.xlsx'
        else:
            file_path ='engage-error_list.xlsx'
        file_path = os.path.join(app.company_dir, file_path)
        
        # engage_downloadメソッドを呼び出してDataFrameを取得
        df_download_data = selenium_en.check_download_file(file_path, error_check)
        if df_download_data is None:
            return  #Noneは何らかのエラーなので、None返ってきた場合は処理を終了。
        
        # ログイン処理
        login_success = selenium_en.login(
            #url='https://en-gage.nｆｆet/company_login/login/',
            url='https://en-gage.net/company/job/?PK',
            login_id=app.engage_id,
            password=app.engage_pass
        )
        if not login_success:
            return
        
        try:
            # ページの読み込みが完了するまで待機
            WebDriverWait(bm.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
            app.log_add("https://en-gage.net/company/job/ にアクセスしました。")
        except TimeoutException:
            app.show_message("", "error", "ページの読み込みに失敗しました。")
            return

        # 記事保存用のデータフレームを初期化
        data = {}
        df_download_data = pd.DataFrame(data, dtype=str)
        df_download_data = bm.convert_columns_to_object(df_download_data)
        df_download_data["■更新フラグ"] = df_download_data["■更新フラグ"].astype(bool)
        new_index = 1 #DFの行数管理用変数を初期化。原稿追加時にこの値を参照して原稿を追加する。

        web_elements_cnt = 0    #現在処理中の原稿Noカウント用
        joblist_page_no = 1 #原稿一覧のページ番号

        x_row = 0  # 原稿一覧テーブルの行要素を初期化
        td_index = 5  # 原稿の公開状況取得用。行の6番目のセルを指定
        css_joblist_next_button = '.page.page--next.md_btn.md_btn--flat' #ページングボタンのXpath    
        cnt_joblist = 1 #原稿の合計数管理用F
        app.log_add(f"{joblist_page_no}ページ目")

        # ループを開始し、特定の要素が存在するかをチェック
        web_element = None
        try:
            web_elements = bm.driver.find_elements(By.CSS_SELECTOR, "a.link.link--edit")

        except NoSuchElementException:
            app.log_add("編集できる原稿が見つかりません。終了します")
            return

        #求人原稿一覧ページのループ
        while web_elements_cnt <= len(web_elements):
            current_url = bm.driver.current_url # 原稿一覧ページのURLを取得。原稿詳細から戻る際に単純なブラウザバックだとログアウトしてしまう場合があるため。
            web_element = web_elements[web_elements_cnt]   #扱いやすいよう要素を別の変数に代入   
            # テーブル全体の値を取得後、Tbobyと行を取得
            table_element = bm.driver.find_element(By.CSS_SELECTOR, ".md_table.md_table--sortable")
            tbody_element = table_element.find_element(By.TAG_NAME, "tbody")
            rows = tbody_element.find_elements(By.TAG_NAME, "tr")
            
            #work_idの値からEngageの原稿編集用のURLをDFに保存する専用の関数
            error_extract_work_id = selenium_en.extract_work_id(web_element, df_download_data, new_index)            
            if error_extract_work_id is None:
                return
            
            #原稿一覧ページの各種データ取得処理
            get_joblistPage_error = selenium_en.get_joblistPage_element(rows, df_download_data, bm, new_index, x_row, td_index)

            #デバッグ用：web_welment_cntの数字を変更すると、指定した原稿からループを開始する　
            # if web_elements_cnt < 48:
            #     web_elements_cnt += 1
            #     continue

            #erro_check引数がNoneの場合は原稿詳細を取得する。
            if error_check is None:
                #原稿の個別ページの取得処理
                df_download_data = selenium_en.get_individual_job(web_element, df_download_data, new_index, bm, joblist_page_no, web_elements_cnt)    

                #現在取得している原稿詳細の内容を、download_data.xlsxに追記する。
            if is_file_open(file_path):
                root = tk.Tk()
                root.withdraw()  # メインウィンドウを表示しない
                root.attributes("-topmost", True)  # 最前面に表示
                messagebox.showerror("エラー", f"ファイル {file_path} は既に開かれています。\n閉じてからOKを押してください。", parent=root)
                root.destroy()
                if is_file_open(file_path):
                    root = tk.Tk()
                    root.withdraw()  # メインウィンドウを表示しない
                    root.attributes("-topmost", True)  # 最前面に表示
                    messagebox.showerror("エラー", f"ファイル {file_path} は既に開かれています。\n処理を中断します。", parent=root)
                    app.log_add(f"ファイル {file_path} は既に開かれているため処理を終了します。")
                    root.destroy()
                    return
                
            pd.set_option('display.max_columns', None)
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df_download_data.to_excel(writer, sheet_name='入力リスト', index=False)
                app.log_add(f"{cnt_joblist}件目の求人を入力マスタのシートに保存しました。")

            #直前の原稿一覧ページへ戻る。
            bm.driver.get(current_url)
            WebDriverWait(bm.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
            web_elements = bm.driver.find_elements(By.CSS_SELECTOR, "a.link.link--edit")    #再度リンクボタン一覧を取得
            
            # 現在の日時を取得
            current_datetime = datetime.now().strftime('%Y/%m/%d %H:%M')

            # engage_updateカラムに現在の日時を入力
            df_download_data.loc[new_index - 1, 'engage_update'] = current_datetime
                
            #次の原稿の編集ボタンをクリック
            x_row += 2  # 奇数の行はロード用のダミーなので、2行ずつカウントアップ
            new_index += 1
            web_elements_cnt += 1
            cnt_joblist += 1

            # except PermissionError:
            #     root = tk.Tk()
            #     root.withdraw()  # メインウィンドウを表示しない
            #     messagebox.showerror("エラー", "ファイルが既に開かれています。")
            #     root.destroy()
            # except Exception as e:
            #     app.log_add(f"データの保存中にエラーが発生しました: {e}")

            #最終行後のページング処理　
            if web_elements_cnt >= len(web_elements) :
                #編集できる原稿が見つからなければ、ページネーションボタンを探し、クリックする。見つからなければ処理を抜ける。
                print(f"{joblist_page_no}ページ目の編集できる原稿が見つからなくなりました")
                try:
                    next_page_element = bm.driver.find_element(By.CSS_SELECTOR, css_joblist_next_button)
                except NoSuchElementException:
                    app.log_add("全ての原稿の取得が完了しました。")
                    break
                
                if next_page_element:
                    #記号のページングボタンをクリック >
                    bm.driver.execute_script("arguments[0].scrollIntoView(true);", next_page_element)
                    bm.driver.execute_script("arguments[0].click();", next_page_element)
                    WebDriverWait(bm.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                    #各種カウント用変数を初期化
                    x_row = 0   #原稿の行管理用変数
                    web_elements_cnt = 0
                    joblist_page_no += 1
                    cnt_joblist += 1
                    app.log_add(f"{joblist_page_no}ページ目")
                try:
                    # 再びページのクリックできる要素を探す
                    web_elements = None
                    web_elements = WebDriverWait(bm.driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.link.link--edit"))
                    )
                    # テーブル全体の値を取得後、Tbobyと行を取得
                    table_element = bm.driver.find_element(By.CSS_SELECTOR, ".md_table.md_table--sortable")
                    tbody_element = table_element.find_element(By.TAG_NAME, "tbody")
                    rows = tbody_element.find_elements(By.TAG_NAME, "tr")
                    
                except Exception as e:
                    app.log_add(f"最終ページです")
                    web_elements = []

        #erro_check引数がNoneの場合は、1回のみでaccount.xlslのループを抜ける。Trueの場合はaccount.xlsxの全ての原稿を読み込む。
        if error_check is None:
            break
        elif error_check is True:
            #エラーチェック用のEXCELファイルを保存する。
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df_filtered = df_download_data[df_download_data['掲載状況'].str.contains('修正', na=False)]
                df_filtered.to_excel(writer, sheet_name='入力リスト', index=False)
                app.log_add(f"{cnt_joblist}件目の求人を入力マスタのシートに保存しました。")
                # ここに必要な処理を記述します

                #全ての会社のDFの値を保存する。
                if 'df_download_total' in locals():
                    df_download_total = pd.concat([df_download_total, df_filtered], ignore_index=True)
                else:
                    df_download_total = df_filtered.copy()
                email_body += f"■{current_company_name}({len(df_filtered)}件)\n"
                if 'df_download_total' in locals():
                    for index, row in df_filtered.iterrows():
                        email_body += f"{row['表示用職種名']},{row['原稿更新URL']}\n"
                email_body += "\n"
 
    if error_check:
        # メールアドレスのリストを取得
        email_addresses = []
        try:
            # mail.csvファイルを読み込む
            df_emails = pd.read_csv('settings/mail.csv')
            # 'email'列からメールアドレスをリストに追加
            email_addresses = df_emails['email'].tolist()
        except Exception as e:
            print(f"メールアドレスの読み込み中にエラーが発生しました: {e}")
        # MailSender クラスのインスタンスを作成
        mailer = MailSender('sv13425.xserver.jp', 587, 'report@recruit-robo.info', '4eWEjri233oqwjDFMK')

        subject = f"【{len(df_download_total)}】件：Engageのエラーチェックが完了しました"

        pc_name = socket.gethostname()
        email_body = f"{app.company_dir}：({len(df_download_total)}件)\n実行PC:{pc_name}\n\n" + email_body
    
        # メールの内容を設定
        mailer.set_message(
            from_addr='report@recruit-robo.info',
            to_addrs=email_addresses,
            subject=subject,
            body= email_body,
            cc=[''],
            bcc=['']
        )

        # メールを送信
        mailer.send()

def engage_error_check(app,e=None):
    """
    Engageのエラーに関する情報をリスト形式で取得し処理する関数。

    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    app.log_add("aa")
    # Engage IDリストが空になるまでループ
    while app.engage_id_list:
        # 現在のEngage ID、パスワード、会社名をリストから取得
        current_engage_id = app.engage_id_list.popleft()
        current_engage_pass = app.engage_pass_list.popleft()
        current_company_name = app.company_dir_list.popleft()
        
        #ログイン情報の未入力チェック
        if not current_engage_id or not current_engage_pass or not current_company_name:
            app.log_add("Engage ID、パスワード、または会社名が空です。次のIDに進みます。")
            continue
        
        # アプリケーションのEngage ID、パスワード、会社ディレクトリを設定
        app.engage_id = current_engage_id
        app.engage_pass = current_engage_pass
        app.company_dir = f"company_data/{current_company_name}/"
        
        # ログイン処理など、必要な処理をここに追加
        app.log_add(f"Engage ID: {current_engage_id}、会社名: {current_company_name}の処理を開始します")
        
        # 例: engage_upload(app)を呼び出す
        engage_upload(app)

def engage_edit_joblist(app,e=None):
    """ Engage-upload.xlsxファイルを用いて原稿のアップデート処理を実施。
        ■更新フラグ カラムが1のレコードを見つけ、原稿のコピーを行い、原稿を修正後、保存ボタンを押して処理を終了する。
    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    
    file_path = os.path.join(app.company_dir, 'engage-download.xlsx')
    login = "update"
    engage_upload(app,e,login,file_path)

def engage_public_private_joblist(app,error_check=None ,e=None):
    """engage-download.xlsxの「掲載状況」を元に、掲載状況を変更する一連の処理
        engage_download関数に引数を渡し処理を分岐する。
        検索条件：■更新フラグ　カラムが1のレコード。
    """
    
    file_path = os.path.join(app.company_dir, 'engage-download.xlsx')
    login = "public_private"

    tp = TextProcessor()
    bm = None
    email_body = ""  # email_bodyを初期化
    page_value = 1  # page_value 1ページ目に初期化
    tr_cnt = 0  # tr_cnt を初期化。非公開処理を実行した際にカウントアップ。ページ再読み込み時は０に戻す

    # カウンタ変数を初期化
    loop_count = 0
    #account.xlsxを基にループ。error_checkがNoneならループは1回のみで終了
    while app.engage_id_list:
        if loop_count > 0 and error_check == None:
            app.log_add("更新がすべて完了しました")
            return

        current_engage_id = app.engage_id_list.popleft()
        current_engage_pass = app.engage_pass_list.popleft()
        current_company_name = app.company_dir_list.popleft()
        app.engage_id = current_engage_id
        app.engage_pass = current_engage_pass
        app.company_dir = f"company_data/{current_company_name}/"
        #各種変数初期化
        new_index = 1 #DFの行数管理用
 
        # 既にBMでブラウザを開いている場合はブラウザを閉じる
        if bm is not None:
            bm.close_existing_browser(app)
        if hasattr(bm, 'driver') and bm.driver is not None:
            try:
                bm.driver.quit()
                app.log_add("既存のブラウザを閉じました。")
            except Exception as e:
                app.log_add(f"既存のブラウザを閉じる際にエラーが発生しました: {e}")
            finally:
                bm.driver = None
        bm = BrowserManager()
        selenium_en = EngageSeleniumAutomation(bm.driver, app)

        file_path ='engage-download.xlsx'
        file_path = os.path.join(app.company_dir, file_path)
        
        # engage_downloadメソッドを呼び出してDataFrameを取得
        df_download_data = load_excel_to_df(file_path)

        # 元のインデックスを保持するために、データフレームにインデックスを追加
        df_download_data.reset_index(inplace=True)
        df_download_data.rename(columns={'index': '元の行番号'}, inplace=True)
        df_download_data['元の行番号'] = pd.to_numeric(df_download_data['元の行番号'], errors='coerce') # 元の行番号を整数化 

        #更新対象のレコードだけ読み込む
        if '■更新フラグ' not in df_download_data.columns:
            app.log_add("エラー: '■更新フラグ' カラムが見つかりません。処理を終了します。")
            return
        df_download_data = df_download_data[df_download_data['■更新フラグ'] == 1]
        if df_download_data is None:
            return  #Noneは何らかのエラーなので、None返ってきた場合は処理を終了。
        
        # Engageログイン処理
        login_success = selenium_en.login(
            url='https://en-gage.net/company/job/?PK',login_id=app.engage_id,password=app.engage_pass)
        if not login_success:
            return
        
        try:
            # ページの読み込みが完了するまで待機
            WebDriverWait(bm.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
            app.log_add("https://en-gage.net/company/job/ にアクセスしました。")
        except TimeoutException:
            app.show_message("", "error", "ページの読み込みに失敗しました。")
            return

        app.log_add(f"{new_index}件目")

        web_element = None
        page_number = 0 #DFのページ番号管理用
        # column_nameを設定
        app.column_name = "掲載状況"

        # engage_updateカラムを初期化
        df_download_data['engage_update'] = pd.NaT
        cnt_df_download_data = len(df_download_data)
        while (new_index ) <=  cnt_df_download_data:
            page_number = int(df_download_data['page番号'].values[new_index - 1])
            row_value = int(df_download_data['原稿行'].values[new_index - 1])

            # 編集する原稿ページへリダイレクト。
            # ただしpage_valueとpage_numberが一致している場合はリダイレクトしない
            if  page_value != page_number:
                redirect_url = f"https://en-gage.net/company/job/?sortColumn=2&sortType=0&page={page_number}"
                bm.driver.get(redirect_url)
                WebDriverWait(bm.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                tr_cnt = 0  # tr_cnt を初期化。非公開処理を実行した際にカウントアップ。ページ再読み込み時は０に戻す
                app.log_add(f"{redirect_url} にリダイレクトしました。")

            # 現在のURLからpage引数の値を取得し変数に代入。値そのものがない場合は1を代入                      
            current_url = bm.driver.current_url
            parsed_url = urlparse(current_url)
            query_params = parse_qs(parsed_url.query)
            page_value = int(query_params.get('page', [1])[0])

            # page_numberとrow_valueを数値型に変更できれば変更する。全角の場合も半角に変換して数値型に変換する
            def convert_to_half_width(s):
 
                return s.translate(str.maketrans('１', '1'))

            # page_numberを数値型に変換
            try:
                page_number = int(convert_to_half_width(str(page_number)))
            except ValueError:
                app.log_add(f"page_numberの変換に失敗しました: {page_number}")
            # row_valueの値を数値型に変換
            try:
                row_value = (int(convert_to_half_width(str(row_value)))) 
            except ValueError:
                app.log_add(f"row_valueの変換に失敗しました: {row_value}")

            # dfの 掲載状況 カラムの値を変数に代入
            posting_status = df_download_data.iloc[new_index - 1]['掲載状況']
            if posting_status == '公開':
                status_value = 1
            elif posting_status == '非公開':
                status_value = 2
            else:
                status_value = None

            #掲載状況のXpathを取得し、掲載状況をアップデート。
            #更新する原稿の掲載行が1行目以外なら、原稿行の値にプラス２をする。Xpath取得は２ずつカウントアップするため。
            if row_value == 0:
                row_value = 1
            else:
                #row_value = (row_value * 2) +1
                row_value = (row_value * 2) - 1
            #原稿の掲載状況のXpathを取得。
            xpath_value_public_private = f'//*[@id="jobIndexTable"]/table/tbody/tr[{row_value}]/td[6]/span/select'

            #//*[@id="jobIndexTable"]/table/tbody/tr[1]/td[6]/span/select

            # 原稿の掲載状況を変更。
            try:
                elemeffnt_value = bm.driver.find_element(By.XPATH, xpath_value_public_private).get_attribute('value')
            except NoSuchElementException:

                #行の値が読み込むたびに偶数、奇数が変わる場合への対策。
                row_value = (row_value) + tr_cnt
                xpath_value_public_private = f'//*[@id="jobIndexTable"]/table/tbody/tr[{row_value}]/td[6]/span/select'
            
            #誤って更新できない原稿を指定した際は次の行へ
            try:
                element_value = bm.driver.find_element(By.XPATH, xpath_value_public_private).get_attribute('text')
            except :
                new_index += 1
                continue

            try:
                selenium_en.fill_web_form('XPATH', xpath_value_public_private, 'select_by_value', status_value)

                # 掲載状況の指定が非公開なら非公開理由のクリック処理を実行
                if status_value == 2:
                    # 非公開理由クリック処理
                    WebDriverWait(bm.driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="stopModalForm"]/div/div/div[1]/dl/dd/div/div[6]/label'))
                    )
                    selenium_en.fill_web_form('XPATH', '//*[@id="stopModalForm"]/div/div/div[1]/dl/dd/div/div[6]/label', 'click')
                    # 送信ボタンクリック
                    selenium_en.fill_web_form('XPATH', '//*[@id="stopModalForm"]/div/div/div[3]/a[2]', 'click')
                    tr_cnt += 1 # 非公開処理を実行するとTRが増加するのでその対策

            except TimeoutException:
                #app.log_add(f"要素 {xpath_value_public_private} はクリックできません。")
                new_index += 1
                continue


            #download.xlsxが開かれているかどうかチェック
            if is_file_open(file_path):
                root = tk.Tk()
                root.withdraw()  # メインウィンドウを表示しない
                root.attributes("-topmost", True)  # 最前面に表示
                messagebox.showerror("エラー", f"ファイル {file_path} は既に開かれています。\n閉じてからOKを押してください。", parent=root)
                root.destroy()
                while is_file_open(file_path):
                    time.sleep(1)  # 1秒待機して再チェック
                    if not is_file_open(file_path):
                        app.log_add(f"ファイル {file_path} は既に開かれているため処理を終了します。")
                        return

            # engage_updateカラムに現在の日時を代入
            df_download_data['engage_update'] = pd.Timestamp.now()

            #現在のDFの行をengage-download.xlsxに保存
            if error_check is None:
                df_download_data.at[new_index, 'engage_update'] = pd.Timestamp.now()  # engage_updateカラムに現在の日時を代入

                # openpyxlを使用して既存のExcelファイルを読み込み
                workbook = load_workbook(file_path)
                sheet = workbook['入力リスト']
                
                engage_update_col_idx = df_download_data.columns.get_loc("engage_update")  # engage_updateカラムの列番号を取得
                # sheet.cell(row=current_rownumber, column=engage_update_col_idx - 1, value=df_download_data['engage_update'])  # Excelシートを更新

                # 特定の行の元の行番号を取得し、Excelシートを更新
                for idx, row in df_download_data.iterrows():
                    original_row = row['元の行番号']
                    if pd.notna(original_row):  # original_row が NaN または NaT でないことを確認
                        original_row = int(original_row) + 2  # 元の行番号に1を加えてExcelの行番号に対応
                        sheet.cell(row=original_row, column=engage_update_col_idx, value=row['engage_update'])  # engage_updateを更新
                        sheet.cell(row=original_row, column=engage_update_col_idx - 1, value=1)  # engageを更新

                    else:
                        print(f"元の行番号が NaN または NaT です。インデックス {idx} の行は更新されません。")


                # 50行目のengage_updateのセルに値を代入
                #sheet.cell(row=50, column=engage_update_col_idx, value=df_download_data['engage_update'])
                # Excelファイルを保存
                workbook.save(file_path)
                app.log_add(f"{cnt_df_download_data}件の求人を{file_path}に保存しました。")

            #掲載状況の終了処理
            app.log_add(f"{app.company_name}の掲載情報の更新が完了しました")
            new_index += 1  # new_indexを元に戻します

        def append_to_email_body(email_body, company_name, df_filtered):
            email_body += f"■{company_name}({len(df_filtered)}件)\n"
            for index, row in df_filtered.iterrows():
                email_body += f"{row['表示用職種名']},{row['原稿更新URL']}\n"
            email_body += "\n"
            return email_body
            
        # メールアドレスのリストを取得
        email_addresses = []
        try:
            # mail.csvファイルを読み込む
            df_emails = pd.read_csv('settings/mail.csv')
            # 'email'列からメールアドレスをリストに追加
            email_addresses = df_emails['email'].tolist()
        except Exception as e:
            print(f"メールアドレスの読み込み中にエラーが発生しました: {e}")
        # MailSender クラスのインスタンスを作成
        mailer = MailSender('sv13425.xserver.jp', 587, 'report@recruit-robo.info', '4eWEjri233oqwjDFMK')

        subject = f"【{cnt_df_download_data}】件：Engageの掲載状況の更新が完了しました"

        pc_name = socket.gethostname()
        email_body = f"{app.company_dir}：({cnt_df_download_data}件)\n実行PC:{pc_name}\n\n" + email_body
    
        # メールの内容を設定
        mailer.set_message(
            from_addr='report@recruit-robo.info',
            to_addrs=email_addresses,
            subject=subject,
            body= email_body,
            cc=[''],
            bcc=['']
        )

        # メールを送信
        mailer.send()
        loop_count += 1
        
    return

def engage_update_manuscript(app,bm,row,df_EngageXpath):
    # この関数は渡されたDFの行の値を基にEngageの原稿を更新する。
    """
    Args:
        app (_type_): アプリケーションのインスタンス
        row (_type_): DataFrameの行データ
    """
    try:
        # 原稿更新処理
        column_name = '雇用形態'
        employment_type = row.get('雇用形態', None)
        radioButton_label_name = "employment_status1"

        xpath_value, action_value = get_xpath_and_action(df_EngageXpath, column_name, row[column_name],app)
        if xpath_value is None or action_value is None:
            app.log_add(f"カラム {column_name} のXPathまたはアクション値が無効です。")


        if employment_type:
            if not bm.set_radio_button("employment_status",radioButton_label_name, employment_type, 'name'):
                app.log_add(f"Failed to set radio button for employment status with value: {employment_type}")
        
        # 例: update_manuscript(app, row)
        
        # 更新処理が成功した場合
        return True
    except Exception as e:
        app.log_add(f"更新処理中にエラーが発生しました: {e}")
        # 更新処理が失敗した場合
        return False
    
def indeed_to_engage(app,e=None):
    """
    IndeedからダウンロードしたJobx.xlsxファイルを、Engageでアップロードするengqage-upload.xlsxへ変換する。        
    _summary_

    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    # 既存の関数を使用してエクセルファイルをロード
    mapping_wb = load_excel_to_df(f'{app.company_dir}indeed_to_engage.xlsx')
    if mapping_wb is False:
        app.log_add("indeed_to_engage.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    jobs_wb = load_excel_to_df(f'{app.company_dir}Jobs.xlsx')
    if jobs_wb is False:
        app.log_add("Jobs.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    engage_wb = load_excel_to_df(f'{app.company_dir}/engage-upload.xlsx')
    if engage_wb is False:
        app.log_add("engage-upload.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    mapping_wb = load_excel_to_df(f'{app.company_dir}indeed_to_engage.xlsx')
    jobs_wb = load_excel_to_df(f'{app.company_dir}Jobs.xlsx')
    engage_wb = load_excel_to_df(f'{app.company_dir}/engage-upload.xlsx')

    # マッピングを辞書に読み込む
    mapping_ws = mapping_wb.active
    mapping = {}
    for row in mapping_ws.iter_rows(min_row=2, values_only=True):
        source_col, target_col = row
        mapping[source_col] = target_col

    # ソースファイルのカラムをターゲットファイルにコピーする
    jobs_ws = jobs_wb.active
    engage_ws = engage_wb.active
    for source_col, target_col in mapping.items():
        source_col_idx = openpyxl.utils.cell.column_index_from_string(source_col)
        target_col_idx = openpyxl.utils.cell.column_index_from_string(target_col)
        
        for row in range(1, jobs_ws.max_row + 1):
            cell_value = jobs_ws.cell(row=row, column=source_col_idx).value
            engage_ws.cell(row=row, column=target_col_idx, value=cell_value)

    # 更新されたターゲットファイルを保存する
    engage_wb.save('/mnt/data/Engage-upload-updated.xlsx')
    app.log_add("Engage-upload-updated.xlsx を保存しました。")
    import openpyxl

    # マッピングファイルを読み込む
    mapping_wb = openpyxl.load_workbook('/mnt/data/indeed_to_engage(確認済み).xlsx')
    mapping_ws = mapping_wb.active

    # マッピングを辞書に読み込む
    mapping = {}
    for row in mapping_ws.iter_rows(min_row=2, values_only=True):
        source_col, target_col = row
        mapping[source_col] = target_col

    # ソースファイル（Jobs.xlsx）を読み込む
    jobs_wb = openpyxl.load_workbook('/mnt/data/Jobs.xlsx')
    jobs_ws = jobs_wb.active

    # ターゲットファイル（Engage-upload.xlsx）を読み込む
    engage_wb = openpyxl.load_workbook('/mnt/data/engage-upload.xlsx')
    engage_ws = engage_wb.active

    # ソースファイルのカラムをターゲットファイルにコピーする
    for source_col, target_col in mapping.items():
        source_col_idx = openpyxl.utils.cell.column_index_from_string(source_col)
        target_col_idx = openpyxl.utils.cell.column_index_from_string(target_col)
        
        for row in range(1, jobs_ws.max_row + 1):
            cell_value = jobs_ws.cell(row=row, column=source_col_idx).value
            engage_ws.cell(row=row, column=target_col_idx, value=cell_value)

    # 更新されたターゲットファイルを保存する
    engage_wb.save('/mnt/data/Engage-upload-updated.xlsx')

    print("カラムのコピーが完了しました。")

def engage_error_check(app,e=None):
    """
    Engageのエラーに関する情報をリスト形式で取得し処理する関数。

    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    app.log_add("aa")
    # Engage IDリストが空になるまでループ
    while app.engage_id_list:
        # 現在のEngage ID、パスワード、会社名をリストから取得
        current_engage_id = app.engage_id_list.popleft()
        current_engage_pass = app.engage_pass_list.popleft()
        current_company_name = app.company_dir_list.popleft()
        
        #ログイン情報の未入力チェック
        if not current_engage_id or not current_engage_pass or not current_company_name:
            app.log_add("Engage ID、パスワード、または会社名が空です。次のIDに進みます。")
            continue
        
        # アプリケーションのEngage ID、パスワード、会社ディレクトリを設定
        app.engage_id = current_engage_id
        app.engage_pass = current_engage_pass
        app.company_dir = f"company_data/{current_company_name}/"
        
        # ログイン処理など、必要な処理をここに追加
        app.log_add(f"Engage ID: {current_engage_id}、会社名: {current_company_name}の処理を開始します")
        
        # 例: engage_upload(app)を呼び出す
        engage_upload(app)

def engage_edit_joblist(app,e=None):
    """ Engage-upload.xlsxファイルを用いて原稿のアップデート処理を実施。
        ■更新フラグ カラムが1のレコードを見つけ、原稿のコピーを行い、原稿を修正後、保存ボタンを押して処理を終了する。
    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    
    file_path = os.path.join(app.company_dir, 'engage-download.xlsx')
    login = "update"
    engage_upload(app,e,login,file_path)

def engage_update_manuscript(app,bm,row,df_EngageXpath):
    # この関数は渡されたDFの行の値を基にEngageの原稿を更新する。
    """
    Args:
        app (_type_): アプリケーションのインスタンス
        row (_type_): DataFrameの行データ
    """
    try:
        # 原稿更新処理
        column_name = '雇用形態'
        employment_type = row.get('雇用形態', None)
        radioButton_label_name = "employment_status1"

        xpath_value, action_value = get_xpath_and_action(df_EngageXpath, column_name, row[column_name],app)
        if xpath_value is None or action_value is None:
            app.log_add(f"カラム {column_name} のXPathまたはアクション値が無効です。")


        if employment_type:
            if not bm.set_radio_button("employment_status",radioButton_label_name, employment_type, 'name'):
                app.log_add(f"Failed to set radio button for employment status with value: {employment_type}")
        
        # 例: update_manuscript(app, row)
        
        # 更新処理が成功した場合
        return True
    except Exception as e:
        app.log_add(f"更新処理中にエラーが発生しました: {e}")
        # 更新処理が失敗した場合
        return False
    
    
def engage_update_manuscript(app,bm,row,df_EngageXpath):
    # この関数は渡されたDFの行の値を基にEngageの原稿を更新する。
    """
    Args:
        app (_type_): アプリケーションのインスタンス
        row (_type_): DataFrameの行データ
    """
    try:
        # 原稿更新処理
        column_name = '雇用形態'
        employment_type = row.get('雇用形態', None)
        radioButton_label_name = "employment_status1"

        xpath_value, action_value = get_xpath_and_action(df_EngageXpath, column_name, row[column_name],app)
        if xpath_value is None or action_value is None:
            app.log_add(f"カラム {column_name} のXPathまたはアクション値が無効です。")


        if employment_type:
            if not bm.set_radio_button("employment_status",radioButton_label_name, employment_type, 'name'):
                app.log_add(f"Failed to set radio button for employment status with value: {employment_type}")
        
        # 例: update_manuscript(app, row)
        
        # 更新処理が成功した場合
        return True
    except Exception as e:
        app.log_add(f"更新処理中にエラーが発生しました: {e}")
        # 更新処理が失敗した場合
        return False
    
def indeed_to_engage(app,e=None):
    """
    IndeedからダウンロードしたJobx.xlsxファイルを、Engageでアップロードするengqage-upload.xlsxへ変換する。        
    _summary_

    Args:
        app (_type_): _description_
        e (_type_, optional): _description_. Defaults to None.
    """
    # 既存の関数を使用してエクセルファイルをロード
    mapping_wb = load_excel_to_df(f'{app.company_dir}indeed_to_engage.xlsx')
    if mapping_wb is False:
        app.log_add("indeed_to_engage.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    jobs_wb = load_excel_to_df(f'{app.company_dir}Jobs.xlsx')
    if jobs_wb is False:
        app.log_add("Jobs.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    engage_wb = load_excel_to_df(f'{app.company_dir}/engage-upload.xlsx')
    if engage_wb is False:
        app.log_add("engage-upload.xlsx の読み込みに失敗しました。処理を中断します。")
        return

    mapping_wb = load_excel_to_df(f'{app.company_dir}indeed_to_engage.xlsx')
    jobs_wb = load_excel_to_df(f'{app.company_dir}Jobs.xlsx')
    engage_wb = load_excel_to_df(f'{app.company_dir}/engage-upload.xlsx')

    # マッピングを辞書に読み込む
    mapping_ws = mapping_wb.active
    mapping = {}
    for row in mapping_ws.iter_rows(min_row=2, values_only=True):
        source_col, target_col = row
        mapping[source_col] = target_col

    # ソースファイルのカラムをターゲットファイルにコピーする
    jobs_ws = jobs_wb.active
    engage_ws = engage_wb.active
    for source_col, target_col in mapping.items():
        source_col_idx = openpyxl.utils.cell.column_index_from_string(source_col)
        target_col_idx = openpyxl.utils.cell.column_index_from_string(target_col)
        
        for row in range(1, jobs_ws.max_row + 1):
            cell_value = jobs_ws.cell(row=row, column=source_col_idx).value
            engage_ws.cell(row=row, column=target_col_idx, value=cell_value)

    # 更新されたターゲットファイルを保存する
    engage_wb.save('/mnt/data/Engage-upload-updated.xlsx')
    app.log_add("Engage-upload-updated.xlsx を保存しました。")
    import openpyxl

    # マッピングファイルを読み込む
    mapping_wb = openpyxl.load_workbook('/mnt/data/indeed_to_engage(確認済み).xlsx')
    mapping_ws = mapping_wb.active

    # マッピングを辞書に読み込む
    mapping = {}
    for row in mapping_ws.iter_rows(min_row=2, values_only=True):
        source_col, target_col = row
        mapping[source_col] = target_col

    # ソースファイル（Jobs.xlsx）を読み込む
    jobs_wb = openpyxl.load_workbook('/mnt/data/Jobs.xlsx')
    jobs_ws = jobs_wb.active

    # ターゲットファイル（Engage-upload.xlsx）を読み込む
    engage_wb = openpyxl.load_workbook('/mnt/data/engage-upload.xlsx')
    engage_ws = engage_wb.active

    # ソースファイルのカラムをターゲットファイルにコピーする
    for source_col, target_col in mapping.items():
        source_col_idx = openpyxl.utils.cell.column_index_from_string(source_col)
        target_col_idx = openpyxl.utils.cell.column_index_from_string(target_col)
        
        for row in range(1, jobs_ws.max_row + 1):
            cell_value = jobs_ws.cell(row=row, column=source_col_idx).value
            engage_ws.cell(row=row, column=target_col_idx, value=cell_value)

    # 更新されたターゲットファイルを保存する
    engage_wb.save('/mnt/data/Engage-upload-updated.xlsx')

    print("カラムのコピーが完了しました。")

    
def engage_import_indeed(app,e=None):
    """Indeed形式のファイルの変換処理。Jobx.xlsxを想定"""
    tp = TextProcessor()
    e = ""
    # ファイルパスの作成
    file_path = os.path.join(app.company_dir, 'Jobs.xlsx')
    # Jobs.xlsxファイルが存在するかチェック
    if not os.path.exists(file_path):
        # ファイルが存在しない場合、エラーダイアログを表示
        ft.AlertDialog(title=ft.Text("Error!"), on_dismiss=lambda e: print(f"ファイルが存在しません。"))
        app.show_message(e,"ERROR",f"ファイルが存在しません")
        return  # 関数を終了
    else:# ファイルが存在する場合、Excelファイルを読み込む
        df = pd.read_excel(file_path)
        app.log_add("読み込みOK")
        ft.AlertDialog(title=ft.Text("Error!"), on_dismiss=lambda e: print(f"ファイルが存在しません。"))
    app.log_add("Jobs.xlsxの内容確認中…")
    

    #Jobs.xlsxファイルの未入力箇所チェック：給与（入力がないと投稿が一切出来ないため）
    # '給与（最低額）'または'給与（最高額）'の列に欠損値がある行を取得
    missing_salary_jobs = df[df['給与（最低額）'].isnull()]

    #給与の上限・下限いずれかが入っていない場合のエラー処理。未入力があればアラートを表示し処理を終了
    if not missing_salary_jobs.empty:
        # 欠損値がある行の数を取得
        num_missing = len(missing_salary_jobs)

        # 欠損値がある行の一覧を取得（インデックスと職種名）
        missing_jobs_list = missing_salary_jobs['職種名'].tolist()

        # 欠損値がある行のインデックスを取得
        missing_jobs_index = missing_salary_jobs.index.tolist()

        # 職種名とインデックスを改行で結合
        missing_jobs_str = "\n\n".join(f"行番号: {index}, 職種名: {name}" for index, name in zip(missing_jobs_index, missing_jobs_list))

        # ユーザーにエラーメッセージを表示
        app.show_message(ft.ElevatedButton(text="Show Message", on_click=app.show_message),"Error",f"給与（最低額）が入っていない職種が{num_missing}件あります。\n{missing_jobs_str}")

        # 処理を終了
        raise SystemExit

    app.log_add("郵便番号作成中…[zipcode]")
    # Jobs.xlsxに'zipcode'列が存在しない場合、新しい列を作成
    if 'zipcode' not in df.columns:
        df['zipcode'] = None

    # 勤務地（都道府県・市区町村・町域）＞郵便番号への変換処理開始。
    # 「XX駅」というパターンとその前のスペースを含めて削除する
    df['勤務地（都道府県・市区町村・町域）'] = df['勤務地（都道府県・市区町村・町域）'].apply(tp.remove_station_name)

    # 'zipcode'列が数字以外の行をフィルタリング
    non_numeric_zipcodes = df[~df['zipcode'].astype(str).str.isdigit()]

    # フィルタリングされた行に対してget_zipcode_from_address関数を適用
    df.loc[non_numeric_zipcodes.index, 'zipcode'] = non_numeric_zipcodes['勤務地（都道府県・市区町村・町域）'].apply(get_zipcode_from_address)

    del non_numeric_zipcodes

    app.log_add("Jobs.xlsxを保存中…")

    #df['zipcode'] = df['勤務地（都道府県・市区町村・町域）'].apply(lambda x: get_zipcode_from_address(x))
    # 結果を新しいExcelファイルに保存
    output_file_path = app.company_dir + '/Jobs.xlsx'
    try:
        # Excelファイルを開く試み
        with open(output_file_path, 'r+') as f:
            df.to_excel(output_file_path, index=False)
    except IOError:
        # ファイルが開いている場合、エラーメッセージを表示して処理を中断
        app.show_message(ft.ElevatedButton(text="Show Message", on_click=app.show_message),"Error",f"Excelファイルが開いています。ファイルを閉じてから再度実行してください。")
        raise SystemExit("Excelファイルが開いているため、処理を中断します。")


    # Jobs.xlsxファイルに書き出し
    output_file_path

    # 勤務地（都道府県・市区町村・町域）＞郵便番号の変換処理ここまで

    app.log_add("保存完了")

    print(file_path)

class CandidateScout:
    def __init__(self, browser_manager, app, max_age, min_age, login_url='https://en-gage.net/company/manage/',logout_url='https://en-gage.net/company_login/auth/logout'):  # コンストラクタを修正
        self.bm = browser_manager
        self.app = app
        self.config = ScoutConfig()  # 設定クラスのインスタンス化
        self.config.update_age_range(min_age, max_age)  # 両方の年齢を更新
        self.scout_count = 0
        self.wait = WebDriverWait(self.bm.driver, 5)
        self.selenium_en = EngageSeleniumAutomation(self.bm.driver, self.app)
        self.logout_url = logout_url  # url変数を取得
        self.login_url = login_url
    def setup_browser_session(self):
        """ブラウザセッションの初期設定"""
        self.app.log_add("ブラウザセッションのセットアップを開始")
        
        try:
            # ログアウト処理
            self.bm.driver.get(self.logout_url)
            WebDriverWait(self.bm.driver, 5).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
                
            )
            # ログイン処理
            login_success = self.selenium_en.login(
                url=self.login_url,
                login_id=self.app.engage_id,
                password=self.app.engage_pass
            )
            
            self.app.log_add(f"ログイン結果: {login_success}")
            
            if not login_success:
                self.app.log_add("ログインに失敗しました")
                return False
                
        except Exception as e:
            self.app.log_add(f"セットアップ中に予期せぬエラー: {str(e)}")
            return False
        return True


    
    def _handle_modals(self):
        """モーダルウィンドウの処理"""
        modal_configs = [
            {
                'xpath': '//*[@id="js_modalNumberA"]/div/div[4]/a',
                'name': 'モーダルA'
            },
            {
                'xpath': '//*[@id="js_modalNumberC"]/div/div[6]/a',
                'name': 'モーダルB'
            },
            {
                'xpath': '//*[@id="js_modalNumberAA"]/div/div[5]/a',
                'name': 'モーダルC'
            }
        ]

        for config in modal_configs:
            modal = None
            try:
                modal = WebDriverWait(self.bm.driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, config['xpath']))
                )
            except Exception as e:
                self.app.log_add(f"{config['name']}処理中にエラー: {str(e)}")
            
            if modal:
                try:
                    modal.click()
                    self.app.log_add(f"{config['name']}を閉じました")
                except Exception as e:
                    self.app.log_add(f"{config['name']}クリック中にエラー: {str(e)}")
            else:
                self.app.log_add(f"{config['name']}は表示されていません")

    def setup_search_criteria(self):
        """地域検索条件の設定"""
        try:
            if not self.config.target_prefecture or self.config.target_prefecture is None or pd.isna(self.config.target_prefecture):
                self.app.log_add("都道府県指定なし")
            else:
                # 都道府県選択
                self.selenium_en.fill_web_form(
                    'XPATH', 
                    '//*[@id="md_select-candidatePrefecture"]', 
                    'select_by_visible_text', 
                    self.config.target_prefecture
                )

                # 絞り込みボタンクリック
                refinement = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[3]/div/div[1]/div[3]/div[1]/div/span[3]/div/a'))
                )
                try:
                    self.app.log_add("要素を表示するためにスクロールを開始")
                    self.bm.driver.execute_script("arguments[0].scrollIntoView();", refinement)
                    # スクロール位置を微調整
                    self.bm.driver.execute_script("window.scrollBy(0, -100);")  # 100ピクセル上にスクロール

                    self.app.log_add("要素が表示されました")
                except Exception as e:
                    self.app.log_add(f"スクロール中にエラー: {str(e)}")
                time.sleep(0.5)
                refinement.click()
                self.app.log_add("検索条件設定完了")
            
            # 初回候補者クリック
            first_candidate = WebDriverWait(self.bm.driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="js_candidateList"]/tr[1]/td[3]/a'))
            )
            self.app.log_add("候補者をクリックする前にスクロールを開始")
            self.bm.driver.execute_script("arguments[0].scrollIntoView();", first_candidate)
            # スクロール位置を微調整
            self.bm.driver.execute_script("window.scrollBy(0, -100);")  # 100ピクセル上にスクロール
            self.app.log_add("候補者が表示されました")
            first_candidate.click()
            
     
            return True
        except Exception as e:
            self.app.log_add(f"絞り込みボタンをクリックできませんでした")
            return False

    def check_candidate_eligibility(self):
        """候補者の適格性チェック"""
        try:
            # 年齢チェック
            age_element = self.wait.until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="js_candidateDetail"]/div[2]/div[2]/div[2]/dl/dd[2]'))
            )
            age_text = age_element.text.replace("歳", "").strip()
            if age_text == "":
                raise ValueError("年齢情報が空です")
                return None
            age = int(age_text)
            
            # キーワードチェック
            
            page_source = self.wait.until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="js_candidateDetail"]/div[2]/div[2]'))
            ).text
            has_keywords = self.config.check_keywords(page_source)

            return {
                'eligible': self.config.validate_age(age) and has_keywords,
                'age': age
            }
        except Exception as e:
            self.app.log_add(f"候補者チェック中にエラー: {str(e)}")
            return None

    def process_candidate(self, eligibility_result):
        """スカウト送信処理実行"""
        try:
            # eligibility_resultの検証
            if 'eligible' not in eligibility_result or 'age' not in eligibility_result:
                self.app.log_add("eligibility_resultに必要なデータがありません")
                return False

            if eligibility_result['eligible']:
                class_name = "js_candidateApproach"
                self.scout_count += 1
                action_message = f"スカウト送信完了: {self.scout_count}件目"
            else:
                try:
                    age = int(eligibility_result['age'])
                except (ValueError, TypeError):
                    self.app.log_add("年齢データが不正です")
                    return False
                action_message = f"条件不適合: {age}歳"
                class_name = "js_candidateDismiss"

            action_button = WebDriverWait(self.bm.driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, class_name)))
            action_button.click()
            time.sleep(random.randint(4, 5))  # インデントを修正
            self.app.log_add(action_message)
            return True
        except Exception as e:
            self.app.log_add("ボタン js_candidateApproachが見つかりませんでした: {e}")                
            try:
                # クラス名modalCommitを検索
                try:
                    modal_commit_button = WebDriverWait(self.bm.driver, 13).until(
                        EC.element_to_be_clickable((By.CLASS_NAME, "md_btn--dismiss"))
                    )
                except Exception:
                    modal_commit_button = WebDriverWait(self.bm.driver, 13).until(
                        EC.element_to_be_clickable((By.CLASS_NAME, "btn--dismiss"))
                    )
                # 見つかった場合、メッセージを表示してクリック
                self.app.log_add("最後の候補者として処理を実行します:{e}")
                modal_commit_button.click()
            except Exception as e:
                error_message = f"候補者処理中にエラー: btn--dismissボタンが見つかりません - {str(e)}"
                current_html = self.bm.driver.page_source
                self.app.log_add(error_message)
                return False 
            return False


def airwork_scout(app, min_age=None, max_age=None, job_title_input=None, prefecture_value=[], e=None, error_check=None):

    if not job_title_input:
        app.show_message(e, "エラー", "職種名を入力してください", box_type="error")
        return
   
   # airwork認証情報の取得と初期化
    filename = "account_airwork.xlsx"
    auth_data = engage_auth_check(app, PROGRAM_VER, filename)
    app.airwork_id, app.airwork_pass, app.company_name_airwork, app.airwork_id_list, app.airwork_pass_list, app.company_airwork_name_list = auth_data

    """スカウト処理のメインエントリーポイント"""
    # デフォルト値の設定
    min_age = min_age or "21"
    max_age = max_age or "60"

    bm = None
    processed_companies = []  # 処理済み企業のリスト

    try:
        # 処理対象のアカウント情報を取得
        if error_check is None:
            # シングルモードの場合は最初の1件のみを処理
            if not app.airwork_id_list:
                app.log_add("処理対象のアカウントが存在しません")
                return
            
            process_list = [(app.airwork_id_list.popleft(), app.airwork_pass_list.popleft(), app.company_airwork_name_list.popleft())]
        else:
            # 全件処理モードの場合は全てのアカウントを処理
            process_list = [(app.airwork_id_list.popleft(), app.airwork_pass_list.popleft(), app.company_airwork_name_list.popleft()) for _ in range(len(app.airwork_id_list))]

        # ログインアカウントごとの処理
        for airwork_id, airwork_pass, company_name_airwork in process_list:
            try:
                # アプリケーション状態の更新
                app.airwork_id = airwork_id
                app.airwork_pass = airwork_pass
                app.airwork_airwork_dir = f"company_data/{company_name_airwork}/"
                processed_companies.append(company_name_airwork)

                # 既存のブラウザを閉じる
                if bm and hasattr(bm, 'driver') and bm.driver:
                    try:
                        bm.driver.quit()
                        app.log_add("既存のブラウザを閉じました")
                    except Exception as e:
                        app.log_add(f"ブラウザを閉じる際にエラー: {str(e)}")

                # ブラウザマネージャーの初期化
                login_url = 'https://ats.rct.airwork.net/airplf/login?agt_association_token=SelfServe'
                bm = BrowserManager()
                scout = AirworkSeleniumAutomation(bm, app, max_age, min_age, job_title_input, login_url)

                # ログイン処理
                if not scout.login():
                    raise Exception("ログイン処理に失敗しました")
                
            except Exception as e:
                app.log_add(f"{company_name_airwork}の処理中にエラー発生: {str(e)}")
                # ブラウザの終了処理
                if bm and hasattr(bm, 'driver') and bm.driver:
                    try:
                        bm.driver.quit()
                        app.log_add(f"{company_name_airwork}のブラウザを終了しました")
                    except Exception as e:
                        app.log_add(f"ブラウザ終了時にエラー: {str(e)}")


        # 基本プロフィール取得
        count = 0  # カウントの初期化
        click_count = 0  # クリック回数の初期化
        while click_count < 10:
            # 候補者ボタンをクリック
            try:
                bm.driver.get("https://ats.rct.airwork.net/candidates")
                WebDriverWait(bm.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                app.log_add("候補者ページを表示")
            except Exception as e:
                app.log_add(f"候補者ページ表示中にエラー: {str(e)}")
                break  # エラー時は処理を抜ける

            # 求人内容を検索
            try:
                job_title_input = job_title_input.replace('\n', '').replace('\r', '')
                job_search_element = WebDriverWait(bm.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, f"//p[normalize-space(text()) = '{job_title_input}']/ancestor::tr"))
                )
                app.log_add(f"求人内容{job_title_input} を見つけました")

                # job_search_element内の候補者を探すボタンをクリック
                candidate_search_button = WebDriverWait(job_search_element, 10).until(
                    EC.element_to_be_clickable((By.XPATH, ".//a[contains(text(), '候補者を探す')]"))
                )
                candidate_search_button.click()
                click_count += 1  # クリック回数をインクリメント
            except Exception as e:
                app.log_add(f"求人タイトルが見つかりませんでした: {str(e)}")
                break  # エラー時は処理を抜ける
            
            # 候補者のソート
            try:
                # セレクトボックスを見つける
                sort_select_box = WebDriverWait(bm.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="candidateList"]/div[2]/select'))
                )
                # セレクトボックス内のオプションを取得
                options = sort_select_box.find_elements(By.TAG_NAME, 'option')
                for option in options:
                    if option.text == "新着順":
                        option.click()  # 新着順を選択
                        app.log_add("新着順を選択しました")
            except Exception as e:
                app.log_add(f"新着順の選択中にエラー: {str(e)}")
            time.sleep(3)

            try:
                # 複数の県を選択
                time.sleep(0.2)
                prefectures = [prefecture_value]
                results = scout.select_multiple_prefectures(bm.driver, prefectures)
                if all(results):
                    app.log_add("全ての県が正常に選択されました")
                else:
                    app.log_add("一部の県の選択に失敗しました")
            except Exception as e:
                app.log_add(f"県の選択中にエラー: {str(e)}")

            # スカウト処理のループ
            while True:
                time.sleep(1)
                count += 1  # カウントアップ
                try:
                    time.sleep(random.uniform(4, 6))
                    
                    app.log_add(f"{count}人目の処理を開始します")
                    elements_info = []
                    
                    # プロフィールラベルとXpathの定義
                    PROFILE_LABELS_AND_XPATHS = {   
                        "年齢": '//*[@id="candidateDetail"]/div[1]/div[1]/p[1]',
                        "年収": '//*[@id="candidateDetail"]/div[1]/div[1]/p[2]',
                        "県": '//*[@id="candidateDetail"]/div[1]/div[1]/p[3]',
                        "学歴": '//*[@id="candidateDetail"]/div[1]/div[1]/p[4]'
                    }

                    age_value = None
                    for label, xpath in PROFILE_LABELS_AND_XPATHS.items():
                        try:
                            element = WebDriverWait(bm.driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, xpath))
                            )
                            elements_info.append(f"{label}を取得しました")
                            if label == "年齢":
                                age_value = element.text  # 年齢の値を取得
                                
                            elif label == "年収":
                                income_value = element.text  # 年収の値を取得
                            elif label == "県":
                                prefecture_value = element.text  # 県の値を取得
                            elif label == "学歴":
                                education_value = element.text  # 学歴の値を取得
                        except Exception as e:
                            #app.log_add(f"{label}の取得中にエラー: {str(e)}")
                            #continue
                            break
                        
                    if elements_info:
                        app.log_add(" | ".join(elements_info))
                    
                    # 年齢のXPATHが取得できない場合、ループを抜ける
                    if age_value is None:
                        try:
                            next_page_button = WebDriverWait(bm.driver, 5).until(
                                EC.element_to_be_clickable((By.CLASS_NAME, 'styles_next__3LCdl'))
                            )
                            next_page_button.click()
                            
                            #ページ読み込みを待機
                            WebDriverWait(bm.driver, 10).until(
                                EC.presence_of_element_located((By.TAG_NAME, 'body'))
                            )
                            app.log_add(f"次のページへ")
                            continue
                        except Exception as e:
                            app.log_add(f"候補者が見つかりません")
                            break  # エラー時は大本のループに戻る

                    # 基本プロフィールからの判別処理
                    try:
                        def click_reject_button():
                            WebDriverWait(bm.driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="candidateDetail"]/div[2]/div[2]/button[2]'))
                            ).click()
                            # DOMを再取得
                            bm.driver.execute_script("""
                                var element = document.evaluate(
                                    '//*[@id="candidateDetail"]',
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {
                                    element.parentNode.removeChild(element);
                                }
                            """)
                            WebDriverWait(bm.driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="candidateDetail"]'))
                            )

                        def log_and_continue(message):
                            app.log_add(message)
                            click_reject_button()

                        # 年齢チェック
                        if not scout.config.validate_age(age_value):
                            log_and_continue(f"年齢NG.{age_value}歳")
                            continue

                        # 県チェック
                        target_prefecture = ""
                        prefecture_check = (prefecture_value == target_prefecture) if target_prefecture and prefecture_value is not None else True
                        if not prefecture_check:
                            log_and_continue(f"所在地NG.{target_prefecture}")
                            continue

                        # 学歴の値を判定
                        education_keywords = ["", "", ""]
                        education_check = any(keyword in education_value for keyword in education_keywords) if education_keywords and education_value is not None else True
                        if not education_check:
                            log_and_continue(f"学歴NG. {education_keywords}")
                            continue

                    except Exception as e:
                        app.log_add(f"基本プロフィールの判別処理中にエラー: {str(e)}")
                        continue

                except Exception as e:
                    app.log_add(f"全ての人材を検索完了: {str(e)}")
                    break  # 大本のループに戻る
                
                try:
                    # アプローチボタンクリック
                    button = WebDriverWait(bm.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="candidateDetail"]/div[2]/div[2]/button[1]'))
                    )
                    bm.driver.execute_script("arguments[0].click();", button)
                    # bodyが表示されるまで待機
                    WebDriverWait(bm.driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, 'body'))
                    )
                    
                    # アプローチを送るボタンクリック
                    button = WebDriverWait(bm.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div[4]/div/div/form/footer/div/section/div[2]/button[2]'))
                    )
                    button.click()
                    
                    # ページの読み込み完了まで待機
                    WebDriverWait(bm.driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, 'body'))
                    )
                    
                    app.log_add(f"{count}人目{age_value}歳:アプローチ成功")

                    
                    # 次の人へ
                    continue
                except Exception as e:
                    app.log_add(f"アプローチボタンのクリック中にエラー: {str(e)}")
                    continue  # エラー時は次のループへ
    except Exception as e:
        app.log_add(f"グローバルエラー: {str(e)}")
    finally:
        # 最終的なクリーンアップ
        if bm and hasattr(bm, 'driver') and bm.driver:
            try:
                bm.driver.quit()
                app.log_add("最終的なブラウザ終了処理完了")
            except Exception as e:
                app.log_add(f"最終的なブラウザ終了時にエラー: {str(e)}")

    # 処理完了のログ
    if processed_companies:
        app.log_add(f"処理完了した企業: {', '.join(processed_companies)}")
    app.log_add("全ての処理が完了しました")

def engage_scout(app, min_age=None, max_age=None, e=None, error_check=None):
    """スカウト処理のメインエントリーポイント"""
    # デフォルト値の設定
    min_age = min_age if min_age is not None else "21"
    max_age = max_age if max_age is not None else "60"

    """スカウト処理のメインエントリーポイント"""
    bm = None
    processed_companies = []  # 処理済み企業のリスト

    try:
        # 処理対象のアカウント情報を取得
        if error_check is None:
            # シングルモードの場合は最初の1件のみを処理
            if not app.engage_id_list:
                app.log_add("処理対象のアカウントが存在しません")
                return
            
            process_list = [(
                app.engage_id_list.popleft(),
                app.engage_pass_list.popleft(),
                app.company_dir_list.popleft()
            )]
        else:
            # 全件処理モードの場合は全てのアカウントを処理
            process_list = [
                (app.engage_id_list.popleft(), app.engage_pass_list.popleft(), app.company_dir_list.popleft())
                for _ in range(len(app.engage_id_list))
            ]

        # ログインアカウントごとの処理
        for engage_id, engage_pass, company_name in process_list:
            try:
                # アプリケーション状態の更新
                app.engage_id = engage_id
                app.engage_pass = engage_pass
                app.company_dir = f"company_data/{company_name}/"
                processed_companies.append(company_name)

                # 既存のブラウザを閉じる
                if bm is not None:
                    if hasattr(bm, 'driver') and bm.driver is not None:
                        try:
                            bm.driver.quit()
                            app.log_add("既存のブラウザを閉じました")
                        except Exception as e:
                            app.log_add(f"ブラウザを閉じる際にエラー: {str(e)}")

                # ブラウザマネージャーの初期化
                bm = BrowserManager()
                scout = CandidateScout(bm, app, max_age, min_age)

                # セッションのセットアップ
                # if not scout.setup_browser_session():
                #     continue
                if not scout.setup_browser_session():
                    raise Exception("ログイン処理に失敗しました")
                
                # モーダルの処理
                scout._handle_modals()

                # 検索条件の設定
                if not scout.setup_search_criteria():
                    continue


                # 候補者処理ループ
                while True:
                    try:
                        # 候補者の適格性チェック
                        eligibility_result = scout.check_candidate_eligibility()
                        if eligibility_result is None:
                            break

                        # 候補者の処理
                        if not scout.process_candidate(eligibility_result):
                            break

                    except Exception as e:
                        app.log_add(f"候補者が見つかりません。")
                        break

                app.log_add(f"{company_name}の処理完了 - 送信数: {scout.scout_count}件")

            except Exception as e:
                    app.log_add(f"{company_name}の処理中にエラー発生: {str(e)}")
            finally:
                # ブラウザの終了処理
                if bm is not None and hasattr(bm, 'driver') and bm.driver is not None:
                    try:
                        bm.driver.quit()
                        app.log_add(f"{company_name}のブラウザを終了しました")
                    except Exception as e:
                        app.log_add(f"ブラウザ終了時にエラー: {str(e)}")

    except Exception as e:
        app.log_add(f"グローバルエラー: {str(e)}")
    finally:
        # 最終的なクリーンアップ
        if bm is not None and hasattr(bm, 'driver') and bm.driver is not None:
            try:
                bm.driver.quit()
                app.log_add("最終的なブラウザ終了処理完了")
            except Exception as e:
                app.log_add(f"最終的なブラウザ終了時にエラー: {str(e)}")

    # 処理完了のログ
    if processed_companies:
        app.log_add(f"処理完了した企業: {', '.join(processed_companies)}")
    app.log_add("全ての処理が完了しました")

class MyApp(ft.Control):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page = page
        
        # 基本的な変数を初期化
        self.selected_index = 0
        
        # その他のUI要素を初期化
        self.tab_content = ft.Column()
        self.input_min_age_value = 21  # 年齢の最小値を数値として初期化
        self.input_age_value = 60  # 年齢の最大値を数値として初期化
        self.job_title_value = None  # 職種名の値を保持する変数を追加
        self.prefecture_value = None  # 都道府県の値を保持する変数を追加
        
        # TextFieldsを初期化（これを先に行う）
        self.text_fields = [ft.TextField(multiline=True, width=400, height=300) for _ in range(3)]
        
        # インスタンス変数として初期化
        self.engage_id = None
        self.engage_pass = None
        self.company_name = None
        self.company_dir = None
        self.engage_id_list = deque()
        self.engage_pass_list = deque()
        self.company_dir_list = deque()

        # 設定関連の変数を初期化
        self.setting_text_over = None
        self.setting_addr_auto_insert_button = None
        self.setting_job_category_not_found = None
        self.setting_debug_mode = None
        
        # プログラムバージョンの定義
        self.PROGRAM_VER = PROGRAM_VER
        
        # NavigationRailの初期設定
        self.navigation_rail = ft.NavigationRail(
            selected_index=self.selected_index,
            on_change=self.on_navigation_rail_change,
            destinations=[
                ft.NavigationRailDestination(
                    icon=ft.icons.HOME,
                    label="Tab 1",
                ),
                ft.NavigationRailDestination(
                    icon=ft.icons.BUSINESS,
                    label="Tab 2",
                ),
                ft.NavigationRailDestination(
                    icon=ft.icons.SCHOOL,
                    label="Tab 3",
                ),
            ],
            width=80,
            extended=False,
            min_width=80,
            min_extended_width=80,
        )
        
        # メインレイアウトの設定
        main_row = ft.Row([
            ft.Container(
                content=self.navigation_rail,
                width=80,
                bgcolor=ft.colors.SURFACE_VARIANT,
            ),
            ft.VerticalDivider(width=1),
            self.tab_content
        ], expand=True)
        
        # ページにメインレイアウトを追加
        self.page.controls.append(main_row)
        
        # 最後にタブコンテンツを表示
        self.show_tab_content()

    def handle_prefecture_change(self, e):
        """都道府県選択の変更を処理するメソッド"""
        self.prefecture_value = e.control.value
        e.control.update()
        
    def handle_job_title_change(self, e):
        """職種名入力の変更を処理するメソッド"""
        self.job_title_value = e.control.value
        e.control.update()
        
    def handle_min_age_change(self, e):
        try:
            value = int(e.control.value or "0")
            if 0 <= value <= 150:  # 妥当な年齢範囲をチェック
                self.input_min_age_value = str(value)
                e.control.error_text = None
            else:
                e.control.error_text = "有効な年齢を入力してください"
                self.input_min_age_value = None
        except ValueError:
            e.control.error_text = "数値を入力してください"
            self.input_min_age_value = None
        e.control.update()

    def handle_max_age_change(self, e):
        """最大年齢入力の変更を処理するメソッド"""
        value = e.control.value
        if value:
            try:
                # 数値として有効かチェック
                self.input_age_value = str(int(value))
                e.control.error_text = None
            except ValueError:
                e.control.error_text = "数値を入力してください"
                self.input_age_value = None
        else:
            self.input_age_value = None
        e.control.update()

    def show_tab_content(self):
        self.tab_content.controls.clear()
        current_text_field = self.text_fields[self.selected_index]
        
        if self.selected_index == 0:
            self.tab_content.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(f"ver {self.PROGRAM_VER}"),
                        ], alignment=ft.MainAxisAlignment.CENTER),
                        ft.Row([
                            ft.ElevatedButton(text="原稿新規アップロード", on_click=self.engage_upload),
                            ft.ElevatedButton(text="原稿ダウンロード", on_click=lambda e: engage_download(self, None, e)),
                            ft.ElevatedButton(text="原稿更新", on_click=lambda e: engage_edit_joblist(self, e)),
                            ft.ElevatedButton(text="原稿公開/非公開", on_click=lambda e: engage_public_private_joblist(self, e)),
                            ft.ElevatedButton(text="エラーチェック", on_click=lambda e: engage_download(self, None, e)),
                        ], alignment=ft.MainAxisAlignment.CENTER, wrap=True),
                        ft.Row([
                            current_text_field
                        ], alignment=ft.alignment.center),
                        ft.Row([
                            ft.Container(
                                content=ft.ElevatedButton(text="管理者にログを送信", 
                                                        on_click=lambda e: self.sendmail_to_admin(e)),
                                alignment=ft.alignment.bottom_right
                            )
                        ], alignment=ft.MainAxisAlignment.END),
                    ], spacing=20),
                    width=800,
                    padding=20,
                    alignment=ft.alignment.center
                )
            )
        elif self.selected_index == 1:
            # 最小年齢の入力フィールド
            min_age_input = ft.TextField(
                label="最小年齢",
                width=50,
                value=self.input_min_age_value,
                on_change=self.handle_min_age_change,
                hint_text="21"
            )

            # 最大年齢の入力フィールド
            max_age_input = ft.TextField(
                label="最大年齢",
                width=50,
                value=self.input_age_value,
                on_change=self.handle_max_age_change,
                hint_text="60"
            )

            # 都道府県のセレクトボックスを追加
            prefecture_input = ft.Dropdown(
                label="都道府県",
                width=100,
                value=self.prefecture_value,  # 現在の値を設定
                on_change=self.handle_prefecture_change,  # 値の変更を処理
                options=[
                    ft.dropdown.Option("北海道"), ft.dropdown.Option("青森県"), ft.dropdown.Option("岩手県"),
                    ft.dropdown.Option("宮城県"), ft.dropdown.Option("秋田県"), ft.dropdown.Option("山形県"),
                    ft.dropdown.Option("福島県"), ft.dropdown.Option("茨城県"), ft.dropdown.Option("栃木県"),
                    ft.dropdown.Option("群馬県"), ft.dropdown.Option("埼玉県"), ft.dropdown.Option("千葉県"),
                    ft.dropdown.Option("東京都"), ft.dropdown.Option("神奈川県"), ft.dropdown.Option("新潟県"),
                    ft.dropdown.Option("富山県"), ft.dropdown.Option("石川県"), ft.dropdown.Option("福井県"),
                    ft.dropdown.Option("山梨県"), ft.dropdown.Option("長野県"), ft.dropdown.Option("岐阜県"),
                    ft.dropdown.Option("静岡県"), ft.dropdown.Option("愛知県"), ft.dropdown.Option("三重県"),
                    ft.dropdown.Option("滋賀県"), ft.dropdown.Option("京都府"), ft.dropdown.Option("大阪府"),
                    ft.dropdown.Option("兵庫県"), ft.dropdown.Option("奈良県"), ft.dropdown.Option("和歌山県"),
                    ft.dropdown.Option("鳥取県"), ft.dropdown.Option("島根県"), ft.dropdown.Option("岡山県"),
                    ft.dropdown.Option("広島県"), ft.dropdown.Option("山口県"), ft.dropdown.Option("徳島県"),
                    ft.dropdown.Option("香川県"), ft.dropdown.Option("愛媛県"), ft.dropdown.Option("高知県"),
                    ft.dropdown.Option("福岡県"), ft.dropdown.Option("佐賀県"), ft.dropdown.Option("長崎県"),
                    ft.dropdown.Option("熊本県"), ft.dropdown.Option("大分県"), ft.dropdown.Option("宮崎県"),
                    ft.dropdown.Option("鹿児島県"), ft.dropdown.Option("沖縄県")
                ]
            )

            # 職種名の入力フィールドを追加
            job_title_input = ft.TextField(
                label="職種名",
                width=300,
                value=self.job_title_value,
                on_change=self.handle_job_title_change,
                hint_text="例：営業職"
            )

            self.tab_content.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("スカウト送信"),
                        ft.Text("条件設定"),
                        ft.Row([
                            min_age_input,
                            ft.Text("〜"),
                            max_age_input,
                            prefecture_input,  # 都道府県セレクトボックス
                            job_title_input  # 職種名入力フィールド
                        ], alignment=ft.alignment.center, spacing=10),
                        ft.Row([
                            ft.ElevatedButton(
                                text="Engage", 
                                on_click=lambda e: engage_scout(
                                    self, 
                                    min_age=self.input_min_age_value,
                                    max_age=self.input_age_value,
                                    e=e
                                )
                            ),
                    ft.ElevatedButton(
                        text="Airwork2.0", 
                        on_click=lambda e: airwork_scout(
                            self, 
                            min_age=self.input_min_age_value,
                            max_age=self.input_age_value,
                            job_title_input=self.job_title_value,  # パラメータ名を正しく修正
                            prefecture_value=self.prefecture_value,  # パラメータ名を正しく修正
                            e=e
                        )
                    ),
                        ], alignment=ft.alignment.center, spacing=10),
                        ft.Row([
                            ft.ElevatedButton(text="indeed → Engage", 
                                            on_click=lambda e: indeed_to_engage(self, e)),
                            ft.ElevatedButton(text="Indeed", 
                                            on_click=lambda e: engage_import_indeed(self, e)),
                        ], alignment=ft.alignment.center, wrap=True),
                        current_text_field,
                    ], spacing=20),
                    width=800,
                    padding=20,
                    alignment=ft.alignment.center
                )
            )
        elif self.selected_index == 2:
            self.tab_content.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Text("Tab 3 content"),
                        current_text_field
                    ], spacing=20),
                    width=800,
                    padding=20,
                    alignment=ft.alignment.center
                )
            )
        
        self.page.update()
    def on_navigation_rail_change(self, e):
        self.selected_index = e.control.selected_index
        self.show_tab_content()

    def log_add(self, text):
        current_text_field = self.text_fields[self.selected_index]
        current_value = current_text_field.value
        new_value = (current_value + "\n" if current_value else "") + text
        current_text_field.value = new_value
        current_text_field.update()

    def engage_upload(self, e):
        self.log_add("原稿新規アップロードが実行されました。")

    def show_message(self, e, title, message, box_type="error"):
        result = None
        event = threading.Event()

        def on_ok(e):
            nonlocal result
            result = True
            self.page.dialog.open = False
            self.page.update()
            event.set()

        def on_yes(e):
            nonlocal result
            result = True
            self.page.dialog.open = False
            self.page.update()
            event.set()

        def on_no(e):
            nonlocal result
            result = False
            self.page.dialog.open = False
            self.page.update()
            event.set()

        if box_type == "error":
            dialog = ft.AlertDialog(
                title=ft.Text(title),
                content=ft.Text(message),
                actions=[
                    ft.TextButton("OK", on_click=on_ok)
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
        elif box_type == "askyesno":
            dialog = ft.AlertDialog(
                title=ft.Text(title),
                content=ft.Text(message),
                actions=[
                    ft.TextButton("Yes", on_click=on_yes),
                    ft.TextButton("No", on_click=on_no)
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()

        event.wait()
        return result

    def initialize_auth_data(self, engage_id, engage_pass, company_name, engage_id_list, engage_pass_list, company_name_list):
        """認証関連のデータを初期化するメソッド"""
        self.engage_id = engage_id
        self.engage_pass = engage_pass
        self.company_name = company_name
        self.engage_id_list = deque(engage_id_list)
        self.engage_pass_list = deque(engage_pass_list)
        self.company_dir_list = deque(company_name_list)
        self.company_dir = f"company_data/{company_name}/"
ft.app(target=main)